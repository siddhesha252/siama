import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import json
import numpy as np
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
import database as db

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="SIAMA Toolbox",
    page_icon="🎨",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
    <style>
    /* ═══════════════════════════════════════════
       SIAMA DESIGN SYSTEM v6
    ═══════════════════════════════════════════ */

    /* ── Hide sidebar collapse arrow label (keyboard_double_ text) ── */
    [data-testid="stSidebarCollapseButton"] { display: none !important; }
    button[kind="headerNoPadding"] { display: none !important; }

    /* ── Sidebar background ── */
    [data-testid="stSidebar"] > div:first-child {
        background-color: #f0f2f6 !important;
        padding-top: 1rem;
    }
    /* Sidebar font family only — no size override so Streamlit slide works */
    [data-testid="stSidebar"] * {
        font-family: "Source Sans Pro", -apple-system, BlinkMacSystemFont,
                     "Segoe UI", Roboto, sans-serif !important;
    }

    /* ── Progress block ── */
    .sp-header {
        font-size: 14px; font-weight: 600; color: #31333f;
        margin-bottom: 6px; padding: 0 4px;
    }
    .sp-overall-row {
        display: flex; align-items: center; gap: 8px;
        padding: 0 4px; margin-bottom: 6px;
    }
    .sp-overall-track {
        flex: 1; height: 6px; background: #dee2e6;
        border-radius: 3px; overflow: hidden;
    }
    .sp-overall-track > span {
        display: block; height: 100%; border-radius: 3px;
        background: linear-gradient(90deg, #1f77b4, #4fa8d8);
        transition: width 0.4s ease;
    }
    .sp-overall-pct {
        font-size: 14px; font-weight: 700; color: #1f77b4;
        flex-shrink: 0; width: 36px; text-align: right;
    }
    .sp-mini-row {
        display: flex; align-items: center; gap: 6px;
        padding: 0 4px; margin-bottom: 4px;
    }
    .sp-mini-row .sp-lbl {
        font-size: 14px; color: #31333f; font-weight: 400;
        width: 32px; flex-shrink: 0;
    }
    .sp-mini-row .sp-track {
        flex: 1; height: 4px; background: #dee2e6;
        border-radius: 2px; overflow: hidden;
    }
    .sp-mini-row .sp-track > span {
        display: block; height: 100%; border-radius: 2px;
        background: #1f77b4;
    }
    .sp-mini-row .sp-track > span.sp-done { background: #28a745; }
    .sp-mini-row .sp-pct {
        font-size: 14px; color: #6c757d; font-weight: 400;
        width: 32px; text-align: right; flex-shrink: 0;
    }

    /* Active project banner */
    .sp-project-banner {
        background: #e8f4fd; border: 1px solid #bee3f8;
        border-radius: 6px; padding: 7px 10px; margin-bottom: 4px;
    }
    .sp-project-banner .sp-pb-label {
        font-size: 11px; font-weight: 600; color: #6c757d;
        text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 3px;
    }
    .sp-project-banner .sp-pb-name {
        font-size: 14px; font-weight: 700; color: #1f77b4;
    }

    

    /* ── Main header (all pages) ── */
    .main-header {
        font-size: 2rem; font-weight: 700; color: #2c5aa0;
        padding: 1rem 1.5rem;
        background: linear-gradient(135deg, #f0f8ff 0%, #e8f4fd 100%);
        border-radius: 8px; margin-bottom: 1.5rem;
        border-left: 4px solid #1f77b4;
    }
    .toolkit-card {
        padding: 1.25rem 1.5rem; border-radius: 8px;
        background-color: #ffffff; border: 1px solid #dee2e6;
        border-left: 4px solid #1f77b4;
        margin: 0.75rem 0; box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    }
    .step-header {
        color: #2c5aa0; font-weight: 600; font-size: 1.1rem;
        margin-top: 0.75rem;
    }

    /* ── Dashboard heading ── */
    .dash-topbar { display: flex; justify-content: space-between; align-items: center; margin-bottom: 1.5rem; }
    .dash-topbar h1 {
        font-size: 2rem; font-weight: 700; color: #2c5aa0;
        padding: 1rem 1.5rem;
        background: linear-gradient(135deg, #f0f8ff 0%, #e8f4fd 100%);
        border-radius: 8px; border-left: 4px solid #1f77b4;
        margin: 0; flex: 1;
    }

    /* ── Dashboard: all text consistent with app body font ── */
    .kpi-card {
        background: #ffffff; border: 1px solid #dee2e6; border-radius: 8px;
        padding: 16px 18px; height: 100%;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        font-family: "Source Sans Pro", -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    .kpi-card.blue   { border-top: 3px solid #1f77b4; }
    .kpi-card.green  { border-top: 3px solid #28a745; }
    .kpi-card.orange { border-top: 3px solid #fd7e14; }
    .kpi-card.purple { border-top: 3px solid #6f42c1; }
    .kpi-card .kpi-label {
        font-size: 11px; font-weight: 600; color: #6c757d;
        text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 8px;
        font-family: inherit;
    }
    .kpi-card .kpi-value {
        font-size: 2.4rem; font-weight: 700; color: #2c5aa0;
        line-height: 1; margin-bottom: 6px; font-family: inherit;
    }
    .kpi-card .kpi-sub { font-size: 12px; color: #6c757d; font-family: inherit; }

    .toolkit-progress-card {
        background: #ffffff; border: 1px solid #dee2e6; border-radius: 8px;
        padding: 14px 16px; margin-bottom: 10px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
        font-family: "Source Sans Pro", -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    .tpc-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 7px; }
    .tpc-header .tpc-name {
        font-size: 14px; font-weight: 700; color: #2c5aa0; font-family: inherit;
    }
    .tpc-header .tpc-count { font-size: 12px; color: #6c757d; font-family: inherit; }
    .tpc-bar { height: 7px; background: #e9ecef; border-radius: 4px; overflow: hidden; margin-bottom: 8px; }
    .tpc-bar > span { display: block; height: 100%; background: linear-gradient(90deg,#1f77b4,#4fa8d8); border-radius:4px; }
    .tpc-bar > span.done { background: linear-gradient(90deg,#28a745,#5cb85c); }
    .tpc-steps { display: flex; gap: 4px; flex-wrap: wrap; }
    .tpc-step-pill {
        padding: 3px 8px; border-radius: 10px; font-size: 11px;
        font-weight: 500; background: #e9ecef; color: #6c757d; font-family: inherit;
    }
    .tpc-step-pill.done { background: #d4edda; color: #155724; }

    .dash-panel {
        background: #ffffff; border: 1px solid #dee2e6; border-radius: 8px;
        padding: 16px 18px; height: 100%; box-shadow: 0 1px 3px rgba(0,0,0,0.04);
        font-family: "Source Sans Pro", -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    .dash-panel h4 {
        font-size: 15px; font-weight: 700; color: #2c5aa0;
        margin: 0 0 4px; font-family: inherit;
    }
    .dash-panel .panel-sub {
        font-size: 12px; color: #6c757d; margin-bottom: 12px;
        padding-bottom: 10px; border-bottom: 1px solid #f0f0f0; font-family: inherit;
    }

    .role-bar-row { display: flex; align-items: center; gap: 10px; padding: 6px 0; border-bottom: 1px solid #f8f9fa; }
    .role-bar-row:last-child { border-bottom: none; }
    .role-bar-row .rbl-name {
        font-size: 13px; font-weight: 500; color: #2c5aa0;
        width: 80px; flex-shrink: 0; font-family: inherit;
    }
    .role-bar-row .rbl-track { flex: 1; height: 7px; background: #e9ecef; border-radius: 3px; overflow: hidden; }
    .role-bar-row .rbl-track > span { display: block; height: 100%; background: linear-gradient(90deg,#1f77b4,#4fa8d8); border-radius: 3px; }
    .role-bar-row .rbl-count {
        font-size: 13px; font-weight: 700; color: #212529;
        width: 20px; text-align: right; font-family: inherit;
    }

    .pri-list-item { display: flex; align-items: center; gap: 10px; padding: 7px 0; border-bottom: 1px solid #f0f0f0; }
    .pri-list-item:last-child { border-bottom: none; }
    .pri-rank {
        width: 24px; height: 24px; border-radius: 50%;
        background: #e8f4fd; border: 1px solid #bee3f8;
        display: flex; align-items: center; justify-content: center;
        font-size: 11px; font-weight: 700; color: #1f77b4; flex-shrink: 0;
        font-family: inherit;
    }
    .pri-info { flex: 1; min-width: 0; }
    .pri-info .pname {
        font-size: 13px; font-weight: 600; color: #212529;
        white-space: nowrap; overflow: hidden; text-overflow: ellipsis; font-family: inherit;
    }
    .pri-info .prole { font-size: 11px; color: #6c757d; margin-top: 1px; font-family: inherit; }
    .pri-score {
        padding: 3px 9px; border-radius: 12px; font-size: 12px;
        font-weight: 700; background: #e8f4fd; color: #1f77b4;
        flex-shrink: 0; font-family: inherit;
    }

    .quad-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; margin-top: 4px; }
    .quad-tile { padding: 12px 14px; border-radius: 6px; border: 1px solid transparent; font-family: inherit; }
    .quad-tile .qt-strategy {
        font-size: 12px; font-weight: 700; text-transform: uppercase;
        letter-spacing: 0.06em; margin-bottom: 4px; font-family: inherit;
    }
    .quad-tile .qt-axis { font-size: 10px; opacity: 0.75; margin-bottom: 8px; font-family: inherit; }
    .quad-tile .qt-count { font-size: 1.8rem; font-weight: 700; line-height: 1; font-family: inherit; }
    .qt-manage { background:#cce5ff; border-color:#b8daff; color:#004085; }
    .qt-satisfy { background:#fff3cd; border-color:#ffeeba; color:#856404; }
    .qt-inform  { background:#d4edda; border-color:#c3e6cb; color:#155724; }
    .qt-monitor { background:#f8f9fa; border-color:#dee2e6; color:#495057; }

    .mat-coverage-grid { display: grid; grid-template-columns: repeat(4,1fr); gap: 8px; margin-top: 6px; }
    .mat-tile {
        padding: 12px 10px; border-radius: 6px; border: 1px solid #dee2e6;
        text-align: center; background: #f8f9fa; font-family: inherit;
    }
    .mat-tile.mat-done { background: #d4edda; border-color: #c3e6cb; }
    .mat-tile .mt-num { font-size: 10px; color: #6c757d; font-weight: 500; font-family: inherit; }
    .mat-tile.mat-done .mt-num { color: #155724; }
    .mat-tile .mt-name {
        font-size: 12px; font-weight: 600; color: #495057;
        margin-top: 4px; font-family: inherit;
    }
    .mat-tile.mat-done .mt-name { color: #155724; }

    .fivep-row { display: flex; align-items: center; gap: 10px; padding: 6px 0; border-bottom: 1px solid #f8f9fa; }
    .fivep-row:last-child { border-bottom: none; }
    .fivep-row .fp-p { font-size: 12px; font-weight: 700; color: #1f77b4; width: 16px; text-align: center; flex-shrink: 0; }
    .fivep-row .fp-name { font-size: 12px; font-weight: 500; color: #495057; width: 75px; flex-shrink: 0; font-family: inherit; }
    .fivep-row .fp-track { flex: 1; height: 5px; background: #e9ecef; border-radius: 3px; overflow: hidden; }
    .fivep-row .fp-track > span { display: block; height: 100%; border-radius: 3px; }
    .fivep-row .fp-track > span.cur { background: #1f77b4; }
    .fivep-row .fp-track > span.des { background: #fd7e14; }
    .fivep-row .fp-cnt { font-size: 11px; font-weight: 600; color: #6c757d; width: 16px; text-align: right; font-family: inherit; }

    .completion-strip {
        display: flex; gap: 0; border: 1px solid #dee2e6;
        border-radius: 8px; overflow: hidden; margin-bottom: 1.25rem;
        font-family: "Source Sans Pro", -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    .cs-item { flex: 1; padding: 10px 14px; border-right: 1px solid #dee2e6; background: #ffffff; text-align: center; }
    .cs-item:last-child { border-right: none; }
    .cs-item.cs-done   { background: #f0fff4; }
    .cs-item.cs-active { background: #e8f4fd; }
    .cs-item .cs-name  {
        font-size: 11px; font-weight: 600; color: #6c757d;
        text-transform: uppercase; letter-spacing: 0.07em; font-family: inherit;
    }
    .cs-item.cs-active .cs-name { color: #1f77b4; }
    .cs-item.cs-done   .cs-name { color: #28a745; }
    .cs-item .cs-count {
        font-size: 13px; font-weight: 700; color: #212529;
        margin-top: 2px; font-family: inherit;
    }
    .cs-item.cs-done .cs-count::before { content: "✓ "; color: #28a745; }

    /* ── Sidebar scroll fix ── */
    [data-testid="stSidebar"] > div:first-child {
        overflow-y: auto !important;
        overflow-x: hidden !important;
        max-height: 100vh !important;
    }
    </style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# SESSION STATE INIT
# ─────────────────────────────────────────────
if 'current_page' not in st.session_state:
    st.session_state.current_page = 'home'
if 'sit_data' not in st.session_state:
    st.session_state.sit_data = {'stakeholders': [], 'roles': {}}
if 'sat_data' not in st.session_state:
    st.session_state.sat_data = {'relationship_matrix': pd.DataFrame()}
if 'mat_data' not in st.session_state:
    st.session_state.mat_data = {}
if 'nature_of_craft' not in st.session_state:
    st.session_state.nature_of_craft = {'current_status': {}, 'desired_status': {}}
if 'current_project' not in st.session_state:
    st.session_state.current_project = None   # project name (string)
if 'current_project_id' not in st.session_state:
    st.session_state.current_project_id = None  # DB row id (int)


# ─────────────────────────────────────────────
# DB HELPER — load a full project into session state
# ─────────────────────────────────────────────
def load_project_from_db(project_id: int, project_name: str):
    """Fetch all data for a project from Supabase and populate session_state."""
    # SIT
    stakeholders = db.get_sit_stakeholders(project_id)
    actors_raw = db.get_sit_actors(project_id)
    roles = {}
    for a in actors_raw:
        r = a.get('role', 'Unknown')
        if r not in roles:
            roles[r] = []
        roles[r].append({
            'name': a.get('name', ''),
            'location': a.get('location', ''),
            'contact': a.get('contact', ''),
            'details': a.get('details', ''),
            '_db_id': a.get('id')
        })
    st.session_state.sit_data = {
        'stakeholders': stakeholders,
        'roles': roles
    }

    # SAT
    ratings = db.get_sat_ratings(project_id)
    conflicts = db.get_sat_conflicts(project_id)
    knowledge = db.get_sat_knowledge(project_id)
    value_maps = db.get_sat_value_maps(project_id)
    subgroups_raw = db.get_sat_subgroups(project_id)

    subgroups = {}
    for sg in subgroups_raw:
        members = [m['stakeholder'] for m in sg.get('sat_subgroup_members', [])]
        subgroups[sg['name']] = {
            'description': sg.get('description', ''),
            'members': members
        }

    knowledge_dict = {}
    for k in knowledge:
        knowledge_dict[k['group_name']] = {
            'knowledge': k.get('knowledge', ''),
            'responsibilities': k.get('responsibilities', ''),
            'skills': k.get('skills', '')
        }

    st.session_state.sat_data = {
        'relationship_matrix': pd.DataFrame(),
        'relationship_data': ratings,
        'conflict_data': conflicts,
        'knowledge_data': knowledge_dict,
        'value_map': value_maps,
        'subgroups': subgroups,
        'subgroup_assignments': {}
    }

    # MAT
    pestel = db.get_mat_pestel(project_id)
    gap = db.get_mat_gap(project_id)
    segments = db.get_mat_segments(project_id)
    personas = db.get_mat_personas(project_id)
    journey = db.get_mat_journey(project_id)
    mystery = db.get_mat_mystery_shopping(project_id)
    complaints = db.get_mat_complaints(project_id)
    brand = db.get_mat_brand_audit(project_id)

    mat = {}
    if pestel: mat['pestel'] = pestel
    if gap: mat['gap'] = gap
    if segments: mat['behavioral_segments'] = segments
    if personas: mat['personas'] = personas
    if journey: mat['customer_journey'] = {row['stage']: row for row in journey}
    if mystery: mat['mystery_shopping'] = mystery
    if complaints: mat['complaints'] = complaints
    if brand: mat['brand_audit'] = brand
    st.session_state.mat_data = mat

    # Nature of Craft
    current_noc = db.get_nature_of_craft(project_id, 'current') or {}
    desired_noc = db.get_nature_of_craft(project_id, 'desired') or {}
    st.session_state.nature_of_craft = {
        'current_status': current_noc,
        'desired_status': desired_noc
    }

    st.session_state.current_project = project_name
    st.session_state.current_project_id = project_id


def reset_state():
    st.session_state.sit_data = {'stakeholders': [], 'roles': {}}
    st.session_state.sat_data = {'relationship_matrix': pd.DataFrame()}
    st.session_state.mat_data = {}
    st.session_state.nature_of_craft = {'current_status': {}, 'desired_status': {}}
    st.session_state.current_project = None
    st.session_state.current_project_id = None


def _project_required():
    """Show a warning if no project is loaded."""
    if not st.session_state.current_project_id:
        st.warning("⚠️ Go to **📁 Projects** and create or load a project first. Your changes will not be saved until you do.")
        return False
    return True


def _pid():
    return st.session_state.current_project_id


# ─────────────────────────────────────────────
# PROGRESS CALCULATION
# Computes step-completion for each toolkit from session_state.
# Returns (steps_done, total_steps, percentage).
# ─────────────────────────────────────────────
def progress_sit():
    """SIT has 4 steps. Step 1 = ≥1 stakeholder questionnaire saved.
    Step 2 = ≥1 actor added. Step 3 + 4 = ≥1 actor exists (visualisations auto-derive)."""
    sit = st.session_state.sit_data
    has_q = len(sit.get('stakeholders', [])) > 0
    actors_count = sum(len(a) for a in sit.get('roles', {}).values())
    has_actors = actors_count > 0
    steps = [has_q, has_actors, has_actors, has_actors]
    done = sum(steps)
    return done, 4, int(done / 4 * 100)


def progress_sat():
    """SAT has 5 steps. Step 1 = ≥1 rating. Step 2 = ≥1 subgroup.
    Step 3 = ≥1 conflict. Step 4 = ≥1 knowledge entry. Step 5 = ≥1 value map."""
    sat = st.session_state.sat_data
    s1 = len(sat.get('relationship_data', []) or []) > 0
    s2 = len(sat.get('subgroups', {}) or {}) > 0
    s3 = len(sat.get('conflict_data', []) or []) > 0
    s4 = len(sat.get('knowledge_data', {}) or {}) > 0
    s5 = len(sat.get('value_map', []) or []) > 0
    steps = [s1, s2, s3, s4, s5]
    done = sum(steps)
    return done, 5, int(done / 5 * 100)


def progress_mat():
    """MAT has 8 tools. Each key in mat_data presence = 1 tool done."""
    mat = st.session_state.mat_data
    tool_keys = ['pestel', 'gap', 'behavioral_segments', 'personas',
                 'customer_journey', 'mystery_shopping', 'complaints', 'brand_audit']
    done = sum(1 for k in tool_keys if mat.get(k))
    return done, 8, int(done / 8 * 100)


def progress_noc():
    """Nature of Craft: 2 status types. Each is 'done' if ≥1 checkbox is True."""
    noc = st.session_state.get('nature_of_craft', {})
    cur = any(v for v in (noc.get('current_status') or {}).values())
    des = any(v for v in (noc.get('desired_status') or {}).values())
    steps = [cur, des]
    done = sum(steps)
    return done, 2, int(done / 2 * 100)


def progress_overall():
    """Overall completion across SIT (4) + SAT (5) + MAT (8) + NoC (2) = 19 steps."""
    d1, t1, _ = progress_sit()
    d2, t2, _ = progress_sat()
    d3, t3, _ = progress_mat()
    d4, t4, _ = progress_noc()
    done = d1 + d2 + d3 + d4
    total = t1 + t2 + t3 + t4
    return done, total, int(done / total * 100) if total else 0


def render_sidebar_progress(label: str, done: int, total: int, pct: int):
    """Renders a mini section progress row inside the overall progress block."""
    done_cls = "done" if pct == 100 else ""
    st.sidebar.markdown(
        f"""<div class="sb-mini-progress">
            <div class="smp-label">{label}</div>
            <div class="smp-track"><span style="width:{pct}%" class="{done_cls}"></span></div>
            <div class="smp-pct">{pct}%</div>
        </div>""",
        unsafe_allow_html=True,
    )


# JSON export helpers (kept for Summary & Export page)
def _json_default(obj):
    if isinstance(obj, pd.DataFrame):
        return {"__dataframe__": True, "data": obj.to_dict(orient="records")}
    if isinstance(obj, (np.integer,)): return int(obj)
    if isinstance(obj, (np.floating,)): return float(obj)
    if isinstance(obj, datetime): return obj.isoformat()
    return str(obj)


# ─────────────────────────────────────────────
# ─────────────────────────────────────────────
# SIDEBAR — light theme, native Streamlit look
# Matches Image 2: palette icon, bright bg,
# radio buttons, clean emoji icons
# ─────────────────────────────────────────────

# Pre-compute progress
sit_done, sit_total, sit_pct = progress_sit()
sat_done, sat_total, sat_pct = progress_sat()
mat_done, mat_total, mat_pct = progress_mat()
noc_done, noc_total, noc_pct = progress_noc()
all_done, all_total, all_pct = progress_overall()

project_name = st.session_state.get("current_project")

# ── Logo + title ──
st.sidebar.markdown("## 🎨 SIAMA Toolbox")

# ── Active project DROPDOWN (clickable switcher) ──
try:
    _all_projects = db.get_all_projects()
except Exception:
    _all_projects = []

if _all_projects:
    _proj_names   = [p["name"] for p in _all_projects]
    _proj_ids     = [p["id"]   for p in _all_projects]

    # Find current index; default to 0 if none loaded
    _cur_id = st.session_state.get("current_project_id")
    _cur_idx = _proj_ids.index(_cur_id) if _cur_id in _proj_ids else 0

    st.sidebar.markdown(
        '<p style="font-size:9px;font-weight:700;color:#6c757d;'
        'text-transform:uppercase;letter-spacing:0.1em;margin-bottom:2px;">Active Project</p>',
        unsafe_allow_html=True,
    )
    _chosen_name = st.sidebar.selectbox(
        "active_project_select",
        _proj_names,
        index=_cur_idx,
        label_visibility="collapsed",
        key="active_project_select",
    )
    # Auto-load if user picked a different project
    _chosen_id = _proj_ids[_proj_names.index(_chosen_name)]
    if _chosen_id != st.session_state.get("current_project_id"):
        with st.spinner(f"Loading {_chosen_name}…"):
            load_project_from_db(_chosen_id, _chosen_name)
        st.rerun()
else:
    st.sidebar.caption("📁 *No projects yet — create one in Projects*")

# ── Navigation ──
st.sidebar.markdown("---")
_page_labels = [
    "🏠  Home",
    "📁  Projects",
    "1️⃣  SIT — Stakeholder Identification",
    "2️⃣  SAT — Stakeholder Analysis",
    "3️⃣  MAT — Market Analysis",
    "🌿  Nature of Craft",
    "📊  Dashboard & Export",
]
_page_keys = ["HOME", "PROJECTS", "SIT", "SAT", "MAT", "NOC", "DASH"]

_selected = st.sidebar.radio(
    "Navigate",
    _page_labels,
    label_visibility="visible",
    key="main_nav_radio",
)
_page = _page_keys[_page_labels.index(_selected)]

# ── Progress — below navigation ──
if st.session_state.get("current_project_id"):
    st.sidebar.markdown("---")
    done_cls_s = "sp-done" if sit_pct==100 else ""
    done_cls_a = "sp-done" if sat_pct==100 else ""
    done_cls_m = "sp-done" if mat_pct==100 else ""
    done_cls_n = "sp-done" if noc_pct==100 else ""
    st.sidebar.markdown(
        f'''<div class="sp-header">Programme Progress</div>
        <div class="sp-overall-row">
            <div class="sp-overall-track"><span style="width:{all_pct}%"></span></div>
            <div class="sp-overall-pct">{all_pct}%</div>
        </div>
        <div class="sp-mini-row">
            <div class="sp-lbl">SIT</div>
            <div class="sp-track"><span style="width:{sit_pct}%" class="{done_cls_s}"></span></div>
            <div class="sp-pct">{sit_pct}%</div>
        </div>
        <div class="sp-mini-row">
            <div class="sp-lbl">SAT</div>
            <div class="sp-track"><span style="width:{sat_pct}%" class="{done_cls_a}"></span></div>
            <div class="sp-pct">{sat_pct}%</div>
        </div>
        <div class="sp-mini-row">
            <div class="sp-lbl">MAT</div>
            <div class="sp-track"><span style="width:{mat_pct}%" class="{done_cls_m}"></span></div>
            <div class="sp-pct">{mat_pct}%</div>
        </div>
        <div class="sp-mini-row">
            <div class="sp-lbl">NoC</div>
            <div class="sp-track"><span style="width:{noc_pct}%" class="{done_cls_n}"></span></div>
            <div class="sp-pct">{noc_pct}%</div>
        </div>''',
        unsafe_allow_html=True,
    )

    # ── Step detail — below progress, plain rows matching sidebar font ──
    _step_map = {
        "SIT": [
            ("Roles & Questionnaires", sit_done >= 1),
            ("Actor Database",         sit_done >= 2),
            ("Role Card",              sit_done >= 2),
            ("Role Map",               sit_done >= 2),
        ],
        "SAT": [
            ("Relationship Matrix",       sat_done >= 1),
            ("Management Tool",           sat_done >= 2),
            ("Conflict Resolution",       sat_done >= 3),
            ("Knowledge & Responsibility",sat_done >= 4),
            ("Value Exchange Map",        sat_done >= 5),
        ],
        "MAT": [
            ("PESTEL",           bool(st.session_state.mat_data.get("pestel"))),
            ("Gap Analysis",     bool(st.session_state.mat_data.get("gap"))),
            ("Segments",         bool(st.session_state.mat_data.get("behavioral_segments"))),
            ("Personas",         bool(st.session_state.mat_data.get("personas"))),
            ("Customer Journey", bool(st.session_state.mat_data.get("customer_journey"))),
            ("Mystery Shopping", bool(st.session_state.mat_data.get("mystery_shopping"))),
            ("Complaints",       bool(st.session_state.mat_data.get("complaints"))),
            ("Brand Audit",      bool(st.session_state.mat_data.get("brand_audit"))),
        ],
        "NOC": [
            ("Current Status",  noc_done >= 1),
            ("Desired Status",  noc_done >= 2),
        ],
        "DASH": [
            ("Programme Overview", all_done > 0),
            ("Export Data",        False),
        ],
    }
    if _page in _step_map:
        st.sidebar.markdown("---")
        steps = _step_map[_page]
        rows_html = "".join(
            f'<div style="display:flex;align-items:center;gap:8px;padding:3px 4px;">'
            f'<div style="width:8px;height:8px;border-radius:50%;flex-shrink:0;'
            f'background:{"#28a745" if done else "#adb5bd"};"></div>'
            f'<span style="font-size:14px;font-weight:400;line-height:1.5;'
            f'color:{"#155724" if done else "#31333f"};">'
            f'{name}</span></div>'
            for name, done in steps
        )
        st.sidebar.markdown(rows_html, unsafe_allow_html=True)

# ─────────────────────────────────────────────
if _page == "HOME":
    st.markdown('<div class="main-header">SIAMA Toolbox</div>', unsafe_allow_html=True)
    st.markdown("### Stakeholder Identification, Analysis, and Market Assessment")

    st.markdown("""
    <div class="toolkit-card">
    <h3>🎯 Purpose</h3>
    The SIAMA toolbox is a comprehensive framework designed to help craft trainers and training organizations
    systematically plan craft education programs by understanding contextually different needs of artisans in India.
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("""
        <div class="toolkit-card">
        <h3>❓ What should we teach?</h3>
        Addressed through:
        <ul>
        <li>Market study components (MAT)</li>
        <li>Artisan's prior skill, knowledge, and aspirations (SIT, SAT)</li>
        </ul>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown("""
        <div class="toolkit-card">
        <h3>👥 Whom should we teach?</h3>
        Addressed through:
        <ul>
        <li>Stakeholder identification (SIT)</li>
        <li>Stakeholder analysis (SAT)</li>
        </ul>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.subheader("📦 The Three Toolkits")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("""
        **1. Stakeholder Identification Toolkit (SIT)**
        - 4 Sequential Steps
        - Identifies craft value chain actors
        - Maps stakeholder relationships
        - Visualizes supply chain
        """)
    with col2:
        st.markdown("""
        **2. Stakeholder Analysis Toolkit (SAT)**
        - 5 Sequential Tools
        - Analyzes power, interest, legitimacy
        - Conflict resolution strategies
        - Knowledge & responsibility mapping
        """)
    with col3:
        st.markdown("""
        **3. Market Analysis Toolkit (MAT)**
        - 8 Analysis Tools
        - Industry understanding
        - Customer insights
        - Brand assessment
        """)

    st.info("👈 Use the sidebar to navigate. Start by creating a project in **📁 Projects**.")


# ─────────────────────────────────────────────
# PROJECTS  (Supabase-backed)
# ─────────────────────────────────────────────
elif _page == "PROJECTS":
    st.markdown('<div class="main-header">Projects</div>', unsafe_allow_html=True)

    st.markdown("""
        <div class="toolkit-card">
        <h3>📁 Manage Projects</h3>
        Each project saves your SIT, SAT, MAT, and Nature of Craft data to the database.
        Create a new project or load an existing one to begin.
        </div>
    """, unsafe_allow_html=True)

    if st.session_state.current_project:
        st.success(f"🟢 Currently editing: **{st.session_state.current_project}**")
    else:
        st.info("⚪ No project loaded. Create one below or load an existing project.")

    st.markdown("---")
    tab_list, tab_create = st.tabs(["📋 My Projects", "➕ New Project"])

    with tab_create:
        st.markdown("### Create a new project")
        new_name = st.text_input("Project name", placeholder="e.g. Bandhani Training Program – Kutch 2026")
        new_desc = st.text_area("Description (optional)", placeholder="Brief description of this training programme")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("💾 Create project", type="primary"):
                name = (new_name or "").strip()
                if not name:
                    st.error("Project name cannot be empty.")
                else:
                    try:
                        created = db.create_project(name, new_desc)
                        if created:
                            reset_state()
                            st.session_state.current_project = name
                            st.session_state.current_project_id = created['id']
                            st.success(f"✅ Project '{name}' created! You can now start filling in SIT, SAT, and MAT.")
                            st.rerun()
                        else:
                            st.error("Failed to create project. Check your Supabase connection.")
                    except Exception as e:
                        st.error(f"Error: {e}")
        with col2:
            if st.button("🆕 Clear current work", help="Clears session data without creating a project"):
                reset_state()
                st.success("Session cleared.")
                st.rerun()

    with tab_list:
        try:
            all_projects = db.get_all_projects()
        except Exception as e:
            st.error(f"Could not fetch projects: {e}")
            all_projects = []

        if not all_projects:
            st.info("No projects yet. Create one in the '➕ New Project' tab.")
        else:
            st.markdown(f"### {len(all_projects)} project(s)")
            for proj in all_projects:
                pid = proj['id']
                pname = proj['name']
                pdesc = proj.get('description', '')
                pcreated = proj.get('created_at', 'N/A')
                is_current = (pid == st.session_state.current_project_id)
                header = ("🟢 " if is_current else "") + f"📁 **{pname}**"

                with st.expander(f"{header} — created {str(pcreated)[:10]}", expanded=is_current):
                    if pdesc:
                        st.caption(pdesc)

                    lc, dc = st.columns([1, 1])
                    with lc:
                        if st.button("📂 Load", key=f"load_{pid}"):
                            with st.spinner("Loading project data..."):
                                load_project_from_db(pid, pname)
                            st.success(f"Loaded '{pname}'.")
                            st.rerun()
                    with dc:
                        confirm = st.checkbox("Confirm delete", key=f"confirm_del_{pid}")
                        if st.button("🗑️ Delete", key=f"del_{pid}", disabled=not confirm):
                            db.delete_project(pid)
                            if st.session_state.current_project_id == pid:
                                reset_state()
                            st.success(f"Deleted '{pname}'.")
                            st.rerun()

    # JSON export (portable backup)
    st.markdown("---")
    st.markdown("### 📤 Export current session as JSON")
    if st.session_state.sit_data.get('roles') or st.session_state.mat_data:
        export_data = {
            'project': st.session_state.current_project,
            'sit_data': st.session_state.sit_data,
            'sat_data': st.session_state.sat_data,
            'mat_data': st.session_state.mat_data,
            'export_date': datetime.now().isoformat()
        }
        st.download_button(
            "📥 Download session as JSON",
            data=json.dumps(export_data, indent=2, default=_json_default),
            file_name=f"siama_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json"
        )
    else:
        st.caption("No data in current session to export.")


# ─────────────────────────────────────────────
# SIT
# ─────────────────────────────────────────────
elif _page == "SIT":
    st.markdown('<div class="main-header">SIT - Stakeholder Identification Toolkit</div>', unsafe_allow_html=True)
    st.markdown("""
    SIT helps understand a craft's value chain by recognising all individuals, groups, and organisations
    that have a direct or indirect impact on it.
    """)

    tab1, tab2, tab3, tab4 = st.tabs([
        "Step 1: Predefined Roles & Questionnaires",
        "Step 2: Role × Actor Database",
        "Step 3: Role Card",
        "Step 4: Role Map"
    ])

    with tab1:
        st.markdown('<p class="step-header">Step 1: Predefined Roles and Questionnaires</p>', unsafe_allow_html=True)
        st.info("Structured questionnaires for each role in the craft value chain.")

        roles_list = ["Supplier", "Producer", "Refiner", "Marketer", "Buyer"]
        selected_role = st.selectbox("Select Role to Interview", roles_list)

        questionnaires = {
            "Supplier": ["Who provides the raw materials?", "Where do they source materials from?", "What is their relationship with producers?"],
            "Producer": ["Who creates the craft products?", "What skills do they possess?", "How long have they been practicing?"],
            "Refiner": ["Who adds value to the basic product?", "What refinement processes are used?", "What expertise do they bring?"],
            "Marketer": ["Who promotes the products?", "What channels do they use?", "What is their reach?"],
            "Buyer": ["Who are the end consumers?", "What are their preferences?", "What price points do they prefer?"]
        }

        responses = {}
        for q in questionnaires[selected_role]:
            responses[q] = st.text_area(q, key=f"q_{selected_role}_{q}")

        st.subheader("Secondary Questionnaire")
        secondary_questions = [
            "How frequently do they interact?",
            "What are the payment terms?",
            "Are there any challenges in this relationship?"
        ]
        for sq in secondary_questions:
            responses[sq] = st.text_area(sq, key=f"sq_{selected_role}_{sq}")

        if st.button("Save Stakeholder Data"):
            entry = {
                "role": selected_role,
                "responses": responses,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            st.session_state.sit_data['stakeholders'].append(entry)
            if st.session_state.current_project_id:
                try:
                    db.save_sit_stakeholder(_pid(), selected_role, responses)
                    st.success(f"✅ Data saved for {selected_role} (database + session)")
                except Exception as e:
                    st.warning(f"Saved to session but DB error: {e}")
            else:
                st.success(f"✅ Data saved for {selected_role} (session only — load a project to persist)")

    with tab2:
        st.markdown('<p class="step-header">Step 2: Role × Actor Database</p>', unsafe_allow_html=True)
        st.info("Record structured information about identified stakeholders.")

        if len(st.session_state.sit_data['stakeholders']) > 0:
            data_for_df = []
            for s in st.session_state.sit_data['stakeholders']:
                row = {"Role": s['role'], "Timestamp": s.get('timestamp', '')}
                for q, a in s['responses'].items():
                    row[q[:30]] = a[:50] if a else ""
                data_for_df.append(row)
            st.dataframe(pd.DataFrame(data_for_df), use_container_width=True)

            with st.expander("🗑️ Manage stakeholder entries"):
                for idx, s in enumerate(list(st.session_state.sit_data['stakeholders'])):
                    c1, c2 = st.columns([5, 1])
                    with c1:
                        st.write(f"**{idx+1}.** {s['role']} — {s.get('timestamp', '')}")
                    with c2:
                        if st.button("🗑️ Delete", key=f"del_sit_sh_{idx}"):
                            st.session_state.sit_data['stakeholders'].pop(idx)
                            st.rerun()

        st.subheader("Add Actor Details")
        col1, col2 = st.columns(2)
        with col1:
            actor_name = st.text_input("Actor Name")
            actor_role = st.selectbox("Role", ["Supplier", "Producer", "Refiner", "Marketer", "Buyer"])
        with col2:
            actor_location = st.text_input("Location")
            actor_contact = st.text_input("Contact Information")
        actor_details = st.text_area("Additional Details")

        if st.button("Add Actor"):
            if actor_name:
                if actor_role not in st.session_state.sit_data['roles']:
                    st.session_state.sit_data['roles'][actor_role] = []
                st.session_state.sit_data['roles'][actor_role].append({
                    "name": actor_name,
                    "location": actor_location,
                    "contact": actor_contact,
                    "details": actor_details
                })
                if st.session_state.current_project_id:
                    try:
                        db.save_sit_actor(_pid(), actor_role, actor_name, actor_location, actor_contact, actor_details)
                        st.success(f"✅ Actor {actor_name} saved to database")
                    except Exception as e:
                        st.warning(f"Saved to session but DB error: {e}")
                else:
                    st.success(f"✅ Actor {actor_name} added (session only — load a project to persist)")
            else:
                st.error("Please enter an actor name.")

    with tab3:
        st.markdown('<p class="step-header">Step 3: Role Card Visualization</p>', unsafe_allow_html=True)
        if st.session_state.sit_data['roles']:
            for role in list(st.session_state.sit_data['roles'].keys()):
                actors = st.session_state.sit_data['roles'][role]
                with st.expander(f"📋 {role} ({len(actors)} actors)"):
                    for i, actor in enumerate(list(actors), 1):
                        c1, c2 = st.columns([6, 1])
                        with c1:
                            st.markdown(f"""
                            **{i}. {actor['name']}**
                            - Location: {actor['location']}
                            - Contact: {actor['contact']}
                            - Details: {actor['details']}
                            """)
                        with c2:
                            if st.button("🗑️", key=f"del_actor_{role}_{i-1}"):
                                if '_db_id' in actor and st.session_state.current_project_id:
                                    try:
                                        db.delete_sit_actor(actor['_db_id'])
                                    except Exception:
                                        pass
                                st.session_state.sit_data['roles'][role].pop(i-1)
                                st.rerun()
                    st.markdown("---")
                    _, rc2 = st.columns([3, 1])
                    with rc2:
                        if st.button(f"🗑️ Delete role '{role}'", key=f"del_role_{role}"):
                            del st.session_state.sit_data['roles'][role]
                            st.rerun()

            role_counts = {role: len(actors) for role, actors in st.session_state.sit_data['roles'].items()}
            fig = px.bar(x=list(role_counts.keys()), y=list(role_counts.values()),
                         labels={'x': 'Role', 'y': 'Number of Actors'},
                         title='Stakeholder Distribution by Role')
            st.plotly_chart(fig, use_container_width=True)

            # Sankey flow
            st.markdown("---")
            st.subheader("🔀 Supply-Chain Flow")
            canonical_order = ["Supplier", "Producer", "Refiner", "Marketer", "Buyer"]
            ordered_roles = [r for r in canonical_order if r in role_counts and role_counts[r] > 0]
            extras = [r for r in role_counts.keys() if r not in canonical_order and role_counts[r] > 0]
            ordered_roles.extend(extras)
            if len(ordered_roles) >= 2:
                node_labels = [f"{r} ({role_counts[r]})" for r in ordered_roles]
                palette = ["#1f77b4", "#2ca02c", "#ff7f0e", "#d62728", "#9467bd", "#8c564b", "#e377c2"]
                node_colors = [palette[i % len(palette)] for i in range(len(ordered_roles))]
                src, tgt, val, link_colors = [], [], [], []
                for i in range(len(ordered_roles) - 1):
                    a, b = ordered_roles[i], ordered_roles[i + 1]
                    flow = min(role_counts[a], role_counts[b])
                    if flow > 0:
                        src.append(i); tgt.append(i + 1); val.append(flow)
                        link_colors.append("rgba(31,119,180,0.35)")
                sankey = go.Figure(go.Sankey(
                    arrangement="snap",
                    node=dict(pad=18, thickness=22, line=dict(color="white", width=1),
                              label=node_labels, color=node_colors),
                    link=dict(source=src, target=tgt, value=val, color=link_colors),
                ))
                sankey.update_layout(height=380, margin=dict(l=10, r=10, t=30, b=10),
                                     title="Actors flowing through the value chain")
                st.plotly_chart(sankey, use_container_width=True)
        else:
            st.warning("⚠️ Please add actors in Step 2 first.")

    with tab4:
        st.markdown('<p class="step-header">Step 4: Role Map</p>', unsafe_allow_html=True)
        if st.session_state.sit_data['roles']:
            roles_data = st.session_state.sit_data['roles']
            canonical_order = ["Supplier", "Producer", "Refiner", "Marketer", "Buyer"]
            present_roles = [r for r in canonical_order if r in roles_data and len(roles_data[r]) > 0]
            extra_roles   = [r for r in roles_data if r not in canonical_order and len(roles_data[r]) > 0]
            all_roles = present_roles + extra_roles

            if not all_roles:
                st.warning("⚠️ No actors added yet. Add actors in Step 2.")
            else:
                # ── Build node positions ──
                # Each role = a column. Actors in that role = rows within column.
                # Role label node sits at top of each column.
                node_x, node_y, node_text, node_color, node_size = [], [], [], [], []
                node_hover = []
                edge_x, edge_y = [], []

                ROLE_COLORS = {
                    "Supplier":  "#1f77b4",
                    "Producer":  "#2ca02c",
                    "Refiner":   "#ff7f0e",
                    "Marketer":  "#d62728",
                    "Buyer":     "#9467bd",
                }
                col_spacing = 2.5

                # ── Role header nodes ──
                role_header_idx = {}
                for col_i, role in enumerate(all_roles):
                    actors = roles_data.get(role, [])
                    x = col_i * col_spacing
                    y = 1.0          # role label sits at top
                    role_header_idx[role] = len(node_x)
                    node_x.append(x); node_y.append(y)
                    node_text.append(f"<b>{role}</b><br>({len(actors)} actors)")
                    node_color.append(ROLE_COLORS.get(role, "#636363"))
                    node_size.append(52)
                    node_hover.append(f"{role}: {len(actors)} actor(s)")

                # ── Actor nodes ──
                actor_indices = {}   # (role, actor_name) → node index
                for col_i, role in enumerate(all_roles):
                    actors = roles_data.get(role, [])
                    for row_i, actor in enumerate(actors):
                        x = col_i * col_spacing
                        # Space actors evenly below the role header
                        y = -0.3 - row_i * 0.6
                        idx = len(node_x)
                        actor_indices[(role, actor['name'])] = idx
                        node_x.append(x); node_y.append(y)
                        node_text.append(actor['name'])
                        node_color.append(ROLE_COLORS.get(role, "#636363"))
                        node_size.append(22)
                        loc = actor.get('location', '')
                        node_hover.append(
                            f"<b>{actor['name']}</b><br>"
                            f"Role: {role}<br>"
                            + (f"Location: {loc}" if loc else "")
                        )
                        # Edge: role header → actor node
                        hdr = role_header_idx[role]
                        edge_x += [node_x[hdr], x, None]
                        edge_y += [node_y[hdr], y, None]

                # ── Flow arrows between consecutive roles (role-to-role) ──
                flow_annotations = []
                for i in range(len(all_roles) - 1):
                    x0 = i * col_spacing + 0.28
                    x1 = (i + 1) * col_spacing - 0.28
                    flow_annotations.append(dict(
                        x=x1, y=1.0, ax=x0, ay=1.0,
                        xref='x', yref='y', axref='x', ayref='y',
                        text='', showarrow=True,
                        arrowhead=2, arrowsize=1.4, arrowwidth=2.5,
                        arrowcolor='#1f77b4',
                    ))

                fig = go.Figure()

                # Draw edges (actor membership lines)
                fig.add_trace(go.Scatter(
                    x=edge_x, y=edge_y,
                    mode='lines',
                    line=dict(color='#dee2e6', width=1.2),
                    hoverinfo='skip',
                    showlegend=False,
                ))

                # Draw actor nodes (smaller)
                actor_x = node_x[len(all_roles):]
                actor_y = node_y[len(all_roles):]
                actor_text = node_text[len(all_roles):]
                actor_color = node_color[len(all_roles):]
                actor_hover = node_hover[len(all_roles):]
                fig.add_trace(go.Scatter(
                    x=actor_x, y=actor_y,
                    mode='markers+text',
                    marker=dict(size=14, color=actor_color,
                                line=dict(color='white', width=1.5), opacity=0.85),
                    text=actor_text,
                    textposition='middle right',
                    textfont=dict(size=11, color='#495057'),
                    hovertext=actor_hover, hoverinfo='text',
                    showlegend=False,
                ))

                # Draw role header nodes (larger, bold)
                fig.add_trace(go.Scatter(
                    x=node_x[:len(all_roles)], y=node_y[:len(all_roles)],
                    mode='markers+text',
                    marker=dict(size=node_size[:len(all_roles)],
                                color=node_color[:len(all_roles)],
                                line=dict(color='white', width=2), opacity=0.95),
                    text=node_text[:len(all_roles)],
                    textposition='bottom center',
                    textfont=dict(size=12, color='#212529'),
                    hovertext=node_hover[:len(all_roles)], hoverinfo='text',
                    showlegend=False,
                ))

                max_col = max(len(all_roles) - 1, 0)
                max_actors = max((len(roles_data.get(r, [])) for r in all_roles), default=0)
                fig.update_layout(
                    title=dict(text="Craft Value Chain — Actor Network", font=dict(size=15, color='#2c5aa0')),
                    showlegend=False,
                    height=max(380, 260 + max_actors * 55),
                    margin=dict(l=20, r=20, t=50, b=60),
                    xaxis=dict(showgrid=False, zeroline=False, showticklabels=False,
                               range=[-0.5, max_col * col_spacing + 0.5]),
                    yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                    plot_bgcolor='#ffffff',
                    paper_bgcolor='rgba(0,0,0,0)',
                    annotations=flow_annotations,
                )
                st.plotly_chart(fig, use_container_width=True)

                # Legend row
                c1, c2, c3 = st.columns(3)
                with c1:
                    st.markdown("→ **Product Flow** — Supplier to Buyer")
                with c2:
                    st.markdown("← **Money Flow** — Buyer to Supplier")
                with c3:
                    st.markdown("↔ **Information Flow** — Bidirectional")

                # Actor count summary
                st.markdown("---")
                st.markdown("**Actors in this map:**")
                cols = st.columns(len(all_roles))
                for col, role in zip(cols, all_roles):
                    color = ROLE_COLORS.get(role, "#636363")
                    actors = roles_data.get(role, [])
                    col.markdown(
                        f'<div style="background:{color}22; border-left:3px solid {color}; '
                        f'padding:8px 12px; border-radius:4px;">'
                        f'<div style="font-size:11px;color:#6c757d;text-transform:uppercase;letter-spacing:.07em">{role}</div>'
                        f'<div style="font-size:20px;font-weight:700;color:{color}">{len(actors)}</div>'
                        + "".join(f'<div style="font-size:11px;color:#495057">• {a["name"]}</div>' for a in actors)
                        + '</div>',
                        unsafe_allow_html=True
                    )
        else:
            st.warning("⚠️ Please add actors in Step 2 to visualize the role map.")


# ─────────────────────────────────────────────
# SAT
# ─────────────────────────────────────────────
elif _page == "SAT":
    st.markdown('<div class="main-header">SAT - Stakeholder Analysis Toolkit</div>', unsafe_allow_html=True)
    st.markdown("SAT evaluates the roles, interests, and influences of identified stakeholders.")

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Step 1: Relationship Matrix",
        "Step 2: Management Tool",
        "Step 3: Conflict Resolution",
        "Step 4: Knowledge & Responsibility",
        "Step 5: Value Exchange Map"
    ])

    with tab1:
        st.markdown('<p class="step-header">Step 1: Stakeholder Relationship Matrix</p>', unsafe_allow_html=True)
        st.info("Rate each stakeholder on Power, Interest, Legitimacy, and Urgency.")

        if st.session_state.sit_data['roles']:
            all_stakeholders = []
            for role, actors in st.session_state.sit_data['roles'].items():
                for actor in actors:
                    all_stakeholders.append(f"{actor['name']} ({role})")

            if all_stakeholders:
                selected_stakeholder = st.selectbox("Select Stakeholder to Rate", all_stakeholders)
                col1, col2 = st.columns(2)
                with col1:
                    power = st.slider("Power (ability to influence outcomes)", 1, 10, 5)
                    interest = st.slider("Interest (level of concern/stake)", 1, 10, 5)
                with col2:
                    legitimacy = st.slider("Legitimacy (validity of involvement)", 1, 10, 5)
                    urgency = st.slider("Urgency (immediacy of demands)", 1, 10, 5)

                interactions = st.text_area("How do they interact with others?")
                tasks = st.text_area("What tasks do they perform?")
                knowledge = st.text_area("What knowledge/skills do they share?")

                if st.button("Save Rating"):
                    if 'relationship_data' not in st.session_state.sat_data:
                        st.session_state.sat_data['relationship_data'] = []
                    st.session_state.sat_data['relationship_data'].append({
                        'stakeholder': selected_stakeholder, 'power': power,
                        'interest': interest, 'legitimacy': legitimacy, 'urgency': urgency,
                        'interactions': interactions, 'tasks': tasks, 'knowledge': knowledge
                    })
                    if st.session_state.current_project_id:
                        try:
                            db.save_sat_rating(_pid(), selected_stakeholder, power, interest,
                                               legitimacy, urgency, interactions, tasks, knowledge)
                            st.success("✅ Rating saved to database")
                        except Exception as e:
                            st.warning(f"Saved to session but DB error: {e}")
                    else:
                        st.success("✅ Rating saved (session only)")

                if 'relationship_data' in st.session_state.sat_data and st.session_state.sat_data['relationship_data']:
                    st.subheader("Current Ratings")
                    st.dataframe(pd.DataFrame(st.session_state.sat_data['relationship_data']), use_container_width=True)

                    with st.expander("🗑️ Manage rating entries"):
                        for idx, r in enumerate(list(st.session_state.sat_data['relationship_data'])):
                            c1, c2 = st.columns([5, 1])
                            with c1:
                                st.write(f"**{idx+1}.** {r.get('stakeholder','?')} — "
                                         f"P:{r.get('power','-')}, I:{r.get('interest','-')}, "
                                         f"L:{r.get('legitimacy','-')}, U:{r.get('urgency','-')}")
                            with c2:
                                if st.button("🗑️ Delete", key=f"del_sat_rat_{idx}"):
                                    removed = st.session_state.sat_data['relationship_data'].pop(idx)
                                    sh_name = removed.get('stakeholder')
                                    for sg in st.session_state.sat_data.get('subgroups', {}).values():
                                        if sh_name in sg.get('members', []):
                                            sg['members'].remove(sh_name)
                                    st.session_state.sat_data.get('subgroup_assignments', {}).pop(sh_name, None)
                                    st.rerun()
        else:
            st.warning("⚠️ Please complete SIT first to identify stakeholders.")

    with tab2:
        st.markdown('<p class="step-header">Step 2: Stakeholder Management Tool</p>', unsafe_allow_html=True)

        if 'relationship_data' in st.session_state.sat_data and st.session_state.sat_data['relationship_data']:
            subtab1, subtab2 = st.tabs(["📊 Stakeholder Analysis", "👥 Subgroup Management"])

            with subtab1:
                df = pd.DataFrame(st.session_state.sat_data['relationship_data'])
                chart_type = st.selectbox("Select Comparison",
                                          ["Power vs Interest", "Power vs Legitimacy", "Power vs Urgency"])
                if chart_type == "Power vs Interest":
                    fig = px.scatter(df, x='power', y='interest', text='stakeholder', title='Power vs Interest Matrix')
                elif chart_type == "Power vs Legitimacy":
                    fig = px.scatter(df, x='power', y='legitimacy', text='stakeholder', title='Power vs Legitimacy Matrix')
                else:
                    fig = px.scatter(df, x='power', y='urgency', text='stakeholder', title='Power vs Urgency Matrix')
                fig.update_traces(textposition='top center')
                fig.update_layout(height=500)
                fig.add_hline(y=5, line_dash="dash", line_color="gray")
                fig.add_vline(x=5, line_dash="dash", line_color="gray")
                st.plotly_chart(fig, use_container_width=True)

                st.subheader("Management Strategies")
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**High Power, High Interest:** Manage Closely — Key Players")
                    st.markdown("**High Power, Low Interest:** Keep Satisfied — Important but passive")
                with col2:
                    st.markdown("**Low Power, High Interest:** Keep Informed — Show consideration")
                    st.markdown("**Low Power, Low Interest:** Monitor — Minimal effort")

                # Stakeholder network graph
                st.markdown("---")
                st.subheader("🕸️ Stakeholder Network")
                rdata = st.session_state.sat_data['relationship_data']
                if rdata:
                    names = [r['stakeholder'] for r in rdata]
                    xs = [float(r.get('power', 0) or 0) for r in rdata]
                    ys = [float(r.get('interest', 0) or 0) for r in rdata]
                    name_to_idx = {n: i for i, n in enumerate(names)}
                    edge_x, edge_y = [], []
                    edge_count = 0
                    for sg in st.session_state.sat_data.get('subgroups', {}).values():
                        members = [m for m in sg.get('members', []) if m in name_to_idx]
                        for i in range(len(members)):
                            for j in range(i + 1, len(members)):
                                a, b = name_to_idx[members[i]], name_to_idx[members[j]]
                                edge_x += [xs[a], xs[b], None]
                                edge_y += [ys[a], ys[b], None]
                                edge_count += 1
                    net = go.Figure()
                    if edge_x:
                        net.add_trace(go.Scatter(x=edge_x, y=edge_y, mode="lines",
                                                 line=dict(width=1, color="rgba(100,100,100,0.35)"),
                                                 hoverinfo="skip", showlegend=False))
                    sizes = [12 + (p * 3.6) for p in xs]
                    net.add_trace(go.Scatter(
                        x=xs, y=ys, mode="markers+text",
                        text=names, textposition="top center",
                        marker=dict(size=sizes, color=ys, colorscale="Viridis",
                                    cmin=1, cmax=10, showscale=True,
                                    colorbar=dict(title="Interest"),
                                    line=dict(width=1, color="white")),
                        hovertemplate="<b>%{text}</b><br>Power: %{x}<br>Interest: %{y}<extra></extra>",
                        showlegend=False,
                    ))
                    net.add_hline(y=5, line_dash="dash", line_color="lightgray")
                    net.add_vline(x=5, line_dash="dash", line_color="lightgray")
                    net.update_layout(height=560,
                                      xaxis=dict(title="Power", range=[0, 11], zeroline=False),
                                      yaxis=dict(title="Interest", range=[0, 11], zeroline=False),
                                      margin=dict(l=10, r=10, t=30, b=10),
                                      title=f"{len(names)} stakeholders · {edge_count} shared-subgroup edge(s)")
                    st.plotly_chart(net, use_container_width=True)

            with subtab2:
                st.subheader("Create and Manage Stakeholder Subgroups")
                if 'subgroups' not in st.session_state.sat_data:
                    st.session_state.sat_data['subgroups'] = {}
                if 'subgroup_assignments' not in st.session_state.sat_data:
                    st.session_state.sat_data['subgroup_assignments'] = {}

                clustering_tabs = st.tabs(["📝 Manual Grouping", "🤖 Automatic Clustering"])

                with clustering_tabs[0]:
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        subgroup_name = st.text_input("Subgroup Name", placeholder="e.g., Local Artisans")
                        subgroup_description = st.text_area("Description")
                    with col2:
                        st.markdown("#### Actions")
                        if st.button("➕ Add Subgroup", use_container_width=True):
                            if subgroup_name:
                                if subgroup_name not in st.session_state.sat_data['subgroups']:
                                    st.session_state.sat_data['subgroups'][subgroup_name] = {
                                        'description': subgroup_description, 'members': []
                                    }
                                    if st.session_state.current_project_id:
                                        try:
                                            db.save_sat_subgroup(_pid(), subgroup_name, subgroup_description)
                                        except Exception:
                                            pass
                                    st.success(f"✅ Subgroup '{subgroup_name}' created!")
                                else:
                                    st.warning("⚠️ Subgroup already exists!")

                    if st.session_state.sat_data['subgroups']:
                        st.markdown("---")
                        st.markdown("#### Assign Stakeholders to Subgroups")
                        col1, col2 = st.columns(2)
                        with col1:
                            stakeholders_for_sg = [item['stakeholder'] for item in st.session_state.sat_data['relationship_data']]
                            selected_sh_for_sg = st.selectbox("Select Stakeholder", stakeholders_for_sg)
                        with col2:
                            sg_options = list(st.session_state.sat_data['subgroups'].keys())
                            selected_sg = st.selectbox("Assign to Subgroup", sg_options)

                        if st.button("Assign Stakeholder to Subgroup"):
                            if selected_sh_for_sg not in st.session_state.sat_data['subgroups'][selected_sg]['members']:
                                st.session_state.sat_data['subgroups'][selected_sg]['members'].append(selected_sh_for_sg)
                                st.session_state.sat_data['subgroup_assignments'][selected_sh_for_sg] = selected_sg
                                if st.session_state.current_project_id:
                                    try:
                                        db.assign_to_subgroup(_pid(), selected_sg, selected_sh_for_sg)
                                    except Exception:
                                        pass
                                st.success(f"✅ {selected_sh_for_sg} assigned to {selected_sg}")

                with clustering_tabs[1]:
                    st.markdown("### Automatic K-Means Clustering")
                    df = pd.DataFrame(st.session_state.sat_data['relationship_data'])
                    if len(df) >= 2:
                        col1, col2 = st.columns(2)
                        with col1:
                            num_clusters = st.slider("Number of Clusters", min_value=2, max_value=min(8, len(df)), value=3)
                        with col2:
                            features_to_use = st.multiselect("Features for Clustering",
                                                              ["power", "interest", "legitimacy", "urgency"],
                                                              default=["power", "interest", "legitimacy", "urgency"])
                        if st.button("🤖 Generate Clusters", type="primary") and len(features_to_use) >= 2:
                            X = df[features_to_use].values
                            scaler = StandardScaler()
                            X_scaled = scaler.fit_transform(X)
                            kmeans = KMeans(n_clusters=num_clusters, random_state=42, n_init=10)
                            df['cluster'] = kmeans.fit_predict(X_scaled)
                            for i in range(num_clusters):
                                cluster_df = df[df['cluster'] == i]
                                avg_power = cluster_df['power'].mean()
                                avg_interest = cluster_df['interest'].mean()
                                if avg_power > 6.5 and avg_interest > 6.5:
                                    name = f"Key Players (Cluster {i+1})"
                                    desc = "High power, high interest — manage closely"
                                elif avg_power > 6.5:
                                    name = f"Keep Satisfied (Cluster {i+1})"
                                    desc = "High power, lower interest — keep satisfied"
                                elif avg_interest > 6.5:
                                    name = f"Keep Informed (Cluster {i+1})"
                                    desc = "Lower power, high interest — keep informed"
                                else:
                                    name = f"Monitor (Cluster {i+1})"
                                    desc = "Lower power, lower interest — monitor"
                                members = cluster_df['stakeholder'].tolist()
                                st.session_state.sat_data['subgroups'][name] = {'description': desc, 'members': members}
                                for m in members:
                                    st.session_state.sat_data['subgroup_assignments'][m] = name
                            st.success(f"✅ Created {num_clusters} subgroups via K-Means")
                            st.rerun()
                    else:
                        st.info("Add at least 2 rated stakeholders to use clustering.")
        else:
            st.warning("⚠️ Please complete Step 1 to rate stakeholders.")

    with tab3:
        st.markdown('<p class="step-header">Step 3: Conflict Resolution Strategy</p>', unsafe_allow_html=True)
        if 'relationship_data' in st.session_state.sat_data and st.session_state.sat_data['relationship_data']:
            stakeholders = [item['stakeholder'] for item in st.session_state.sat_data['relationship_data']]
            selected_sh = st.selectbox("Select Stakeholder for Conflict Analysis", stakeholders)
            cooperativeness = st.slider("Cooperativeness Level", 1, 10, 5)
            competitiveness = st.slider("Competitiveness Level", 1, 10, 5)
            conflict_description = st.text_area("Describe potential conflicts")

            if st.button("Add Conflict Strategy"):
                if 'conflict_data' not in st.session_state.sat_data:
                    st.session_state.sat_data['conflict_data'] = []
                st.session_state.sat_data['conflict_data'].append({
                    'stakeholder': selected_sh, 'cooperativeness': cooperativeness,
                    'competitiveness': competitiveness, 'description': conflict_description
                })
                if st.session_state.current_project_id:
                    try:
                        db.save_sat_conflict(_pid(), selected_sh, cooperativeness, competitiveness, conflict_description)
                        st.success("✅ Conflict strategy saved to database")
                    except Exception as e:
                        st.warning(f"Saved to session but DB error: {e}")
                else:
                    st.success("✅ Conflict strategy saved (session only)")

            if 'conflict_data' in st.session_state.sat_data and st.session_state.sat_data['conflict_data']:
                df_conflict = pd.DataFrame(st.session_state.sat_data['conflict_data'])
                fig = px.scatter(df_conflict, x='competitiveness', y='cooperativeness',
                                 text='stakeholder', title='Conflict Resolution Matrix')
                fig.update_traces(textposition='top center')
                fig.add_hline(y=5, line_dash="dash", line_color="gray")
                fig.add_vline(x=5, line_dash="dash", line_color="gray")
                st.plotly_chart(fig, use_container_width=True)
        else:
            st.warning("⚠️ Please complete Step 1 first.")

    with tab4:
        st.markdown('<p class="step-header">Step 4: Knowledge and Responsibility Chart</p>', unsafe_allow_html=True)
        if 'relationship_data' in st.session_state.sat_data and st.session_state.sat_data['relationship_data']:
            stakeholder_groups = st.multiselect(
                "Select Stakeholder Group",
                ["Key Players", "Keep Satisfied", "Keep Informed", "Monitor"],
                default=["Key Players"]
            )
            for group in stakeholder_groups:
                with st.expander(f"📚 {group}"):
                    knowledge_areas = st.text_area(f"Knowledge Areas for {group}", key=f"know_{group}")
                    responsibilities = st.text_area(f"Responsibilities for {group}", key=f"resp_{group}")
                    skills_needed = st.text_area(f"Skills Needed for {group}", key=f"skill_{group}")
                    if st.button(f"Save {group} Data", key=f"btn_{group}"):
                        if 'knowledge_data' not in st.session_state.sat_data:
                            st.session_state.sat_data['knowledge_data'] = {}
                        st.session_state.sat_data['knowledge_data'][group] = {
                            'knowledge': knowledge_areas,
                            'responsibilities': responsibilities,
                            'skills': skills_needed
                        }
                        if st.session_state.current_project_id:
                            try:
                                db.save_sat_knowledge(_pid(), group, knowledge_areas, responsibilities, skills_needed)
                                st.success(f"✅ Data saved for {group} to database")
                            except Exception as e:
                                st.warning(f"Saved to session but DB error: {e}")
                        else:
                            st.success(f"✅ Data saved for {group} (session only)")

            if 'knowledge_data' in st.session_state.sat_data:
                st.subheader("Knowledge & Responsibility Summary")
                for group, data in st.session_state.sat_data['knowledge_data'].items():
                    st.markdown(f"**{group}:**")
                    st.write(f"- Knowledge: {data['knowledge']}")
                    st.write(f"- Responsibilities: {data['responsibilities']}")
                    st.write(f"- Skills: {data['skills']}")
        else:
            st.warning("⚠️ Please complete Step 1 first.")

    with tab5:
        st.markdown('<p class="step-header">Step 5: Value Exchange Map</p>', unsafe_allow_html=True)
        if 'relationship_data' in st.session_state.sat_data and st.session_state.sat_data['relationship_data']:
            stakeholder = st.selectbox("Select Stakeholder for Value Mapping",
                                       [item['stakeholder'] for item in st.session_state.sat_data['relationship_data']])
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("### Customer Profile")
                pains = st.text_area("Pains (problems, challenges)")
                gains = st.text_area("Gains (desired outcomes, benefits)")
                jobs = st.text_area("Jobs to be Done")
            with col2:
                st.markdown("### Value Map")
                pain_relievers = st.text_area("Pain Relievers (how training helps)")
                gain_creators = st.text_area("Gain Creators (benefits provided)")
                products_services = st.text_area("Training Products & Services")

            if st.button("Save Value Map"):
                if 'value_map' not in st.session_state.sat_data:
                    st.session_state.sat_data['value_map'] = []
                st.session_state.sat_data['value_map'].append({
                    'stakeholder': stakeholder, 'pains': pains, 'gains': gains, 'jobs': jobs,
                    'pain_relievers': pain_relievers, 'gain_creators': gain_creators,
                    'products_services': products_services
                })
                if st.session_state.current_project_id:
                    try:
                        db.save_sat_value_map(_pid(), stakeholder, pains, gains, jobs,
                                              pain_relievers, gain_creators, products_services)
                        st.success("✅ Value map saved to database")
                    except Exception as e:
                        st.warning(f"Saved to session but DB error: {e}")
                else:
                    st.success("✅ Value map saved (session only)")

            if 'value_map' in st.session_state.sat_data and st.session_state.sat_data['value_map']:
                st.subheader("Saved Value Maps")
                for vm_idx, vm in enumerate(list(st.session_state.sat_data['value_map'])):
                    with st.expander(f"📊 {vm['stakeholder']}"):
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write(f"**Pains:** {vm['pains']}")
                            st.write(f"**Gains:** {vm['gains']}")
                            st.write(f"**Jobs:** {vm['jobs']}")
                        with col2:
                            st.write(f"**Pain Relievers:** {vm['pain_relievers']}")
                            st.write(f"**Gain Creators:** {vm['gain_creators']}")
                            st.write(f"**Products/Services:** {vm['products_services']}")
        else:
            st.warning("⚠️ Please complete Step 1 first.")


# ─────────────────────────────────────────────
# MAT
# ─────────────────────────────────────────────
elif _page == "MAT":
    st.markdown('<div class="main-header">MAT - Market Analysis Toolkit</div>', unsafe_allow_html=True)
    st.markdown("MAT ensures training programs align with market demands through comprehensive analysis.")

    mat_tools = st.selectbox("Select Analysis Tool", [
        "PESTEL Analysis", "Gap Analysis", "Behavioral Segmentation",
        "User Persona", "Customer Journey Map", "Mystery Shopping",
        "Complaint Data Analysis", "Brand Audit"
    ])

    if mat_tools == "PESTEL Analysis":
        st.subheader("PESTEL Analysis")
        col1, col2 = st.columns(2)
        with col1:
            political = st.text_area("Political Factors", placeholder="e.g., PM Vishwakarma scheme")
            economic = st.text_area("Economic Factors", placeholder="e.g., Rising demand for handmade products")
            social = st.text_area("Social Factors", placeholder="e.g., Growing appreciation for traditional crafts")
        with col2:
            technological = st.text_area("Technological Factors", placeholder="e.g., E-commerce enabling direct sales")
            environmental = st.text_area("Environmental Factors", placeholder="e.g., Preference for eco-friendly products")
            legal = st.text_area("Legal Factors", placeholder="e.g., GI tag protection")

        if st.button("Save PESTEL Analysis"):
            data = {'political': political, 'economic': economic, 'social': social,
                    'technological': technological, 'environmental': environmental, 'legal': legal}
            st.session_state.mat_data['pestel'] = data
            if st.session_state.current_project_id:
                try:
                    db.save_mat_pestel(_pid(), data)
                    st.success("✅ PESTEL saved to database")
                except Exception as e:
                    st.warning(f"Saved to session but DB error: {e}")
            else:
                st.success("✅ PESTEL saved (session only)")

    elif mat_tools == "Gap Analysis":
        st.subheader("Gap Analysis")
        col1, col2 = st.columns(2)
        with col1:
            current_state = st.text_area("Current State")
            current_strengths = st.text_area("Current Strengths")
            current_weaknesses = st.text_area("Current Weaknesses")
        with col2:
            desired_state = st.text_area("Desired Future State")
            opportunities = st.text_area("Opportunities")
            threats = st.text_area("Threats")
        action_plan = st.text_area("Action Plan to Bridge the Gap")

        if st.button("Save Gap Analysis"):
            data = {'current_state': current_state, 'current_strengths': current_strengths,
                    'current_weaknesses': current_weaknesses, 'desired_state': desired_state,
                    'opportunities': opportunities, 'threats': threats, 'action_plan': action_plan}
            st.session_state.mat_data['gap'] = data
            if st.session_state.current_project_id:
                try:
                    db.save_mat_gap(_pid(), data)
                    st.success("✅ Gap Analysis saved to database")
                except Exception as e:
                    st.warning(f"Saved to session but DB error: {e}")
            else:
                st.success("✅ Gap Analysis saved (session only)")

    elif mat_tools == "Behavioral Segmentation":
        st.subheader("Behavioral Segmentation")
        segment_name = st.text_input("Segment Name")
        col1, col2 = st.columns(2)
        with col1:
            purchase_behavior = st.text_area("Purchase Behavior")
            usage_rate = st.text_area("Usage Rate")
        with col2:
            benefits_sought = st.text_area("Benefits Sought")
            loyalty_status = st.text_area("Loyalty Status")
        occasion = st.text_area("Purchase Occasion")

        if st.button("Add Segment"):
            if 'behavioral_segments' not in st.session_state.mat_data:
                st.session_state.mat_data['behavioral_segments'] = []
            seg = {'name': segment_name, 'purchase_behavior': purchase_behavior,
                   'usage_rate': usage_rate, 'benefits_sought': benefits_sought,
                   'loyalty_status': loyalty_status, 'occasion': occasion}
            st.session_state.mat_data['behavioral_segments'].append(seg)
            if st.session_state.current_project_id:
                try:
                    db.save_mat_segment(_pid(), seg)
                    st.success(f"✅ Segment '{segment_name}' saved to database")
                except Exception as e:
                    st.warning(f"Saved to session but DB error: {e}")
            else:
                st.success(f"✅ Segment '{segment_name}' added (session only)")

        if 'behavioral_segments' in st.session_state.mat_data:
            for seg_idx, seg in enumerate(list(st.session_state.mat_data['behavioral_segments'])):
                with st.expander(f"👥 {seg['name']}"):
                    st.write(f"Purchase Behavior: {seg['purchase_behavior']}")
                    st.write(f"Benefits Sought: {seg['benefits_sought']}")
                    if st.button("🗑️ Delete", key=f"del_seg_{seg_idx}"):
                        st.session_state.mat_data['behavioral_segments'].pop(seg_idx)
                        st.rerun()

    elif mat_tools == "User Persona":
        st.subheader("User Persona")
        persona_name = st.text_input("Persona Name")
        col1, col2 = st.columns(2)
        with col1:
            age_range = st.text_input("Age Range")
            occupation = st.text_input("Occupation")
            location = st.text_input("Location")
            income_level = st.text_input("Income Level")
        with col2:
            education = st.text_input("Education Level")
            family_status = st.text_input("Family Status")
            lifestyle = st.text_input("Lifestyle")
        goals = st.text_area("Goals & Motivations")
        pain_points = st.text_area("Pain Points & Frustrations")
        shopping_habits = st.text_area("Shopping Habits")

        if st.button("Save Persona"):
            if 'personas' not in st.session_state.mat_data:
                st.session_state.mat_data['personas'] = []
            persona = {'name': persona_name, 'age_range': age_range, 'occupation': occupation,
                       'location': location, 'income_level': income_level, 'education': education,
                       'family_status': family_status, 'lifestyle': lifestyle,
                       'goals': goals, 'pain_points': pain_points, 'shopping_habits': shopping_habits}
            st.session_state.mat_data['personas'].append(persona)
            if st.session_state.current_project_id:
                try:
                    db.save_mat_persona(_pid(), persona)
                    st.success(f"✅ Persona '{persona_name}' saved to database")
                except Exception as e:
                    st.warning(f"Saved to session but DB error: {e}")
            else:
                st.success(f"✅ Persona '{persona_name}' created (session only)")

        if 'personas' in st.session_state.mat_data:
            for p_idx, persona in enumerate(list(st.session_state.mat_data['personas'])):
                with st.expander(f"👤 {persona['name']}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"Age: {persona['age_range']} | Occupation: {persona['occupation']}")
                        st.write(f"Location: {persona['location']} | Income: {persona['income_level']}")
                    with col2:
                        st.write(f"Education: {persona['education']} | Family: {persona['family_status']}")
                    st.write(f"**Goals:** {persona['goals']}")
                    st.write(f"**Pain Points:** {persona['pain_points']}")
                    if st.button("🗑️ Delete", key=f"del_persona_{p_idx}"):
                        st.session_state.mat_data['personas'].pop(p_idx)
                        st.rerun()

    elif mat_tools == "Customer Journey Map":
        st.subheader("Customer Journey Map")
        stages = ["Awareness", "Consideration", "Purchase", "Usage", "Loyalty"]
        journey_data = {}
        for stage in stages:
            with st.expander(f"📍 {stage}"):
                journey_data[stage] = {
                    'touchpoints': st.text_area(f"Touchpoints in {stage}", key=f"tp_{stage}"),
                    'actions': st.text_area("Customer Actions", key=f"act_{stage}"),
                    'emotions': st.text_area("Emotions/Thoughts", key=f"emo_{stage}"),
                    'pain_points': st.text_area("Pain Points", key=f"pain_{stage}"),
                    'opportunities': st.text_area("Opportunities", key=f"opp_{stage}")
                }
        if st.button("Save Customer Journey"):
            st.session_state.mat_data['customer_journey'] = journey_data
            if st.session_state.current_project_id:
                try:
                    db.save_mat_journey(_pid(), journey_data)
                    st.success("✅ Customer Journey saved to database")
                except Exception as e:
                    st.warning(f"Saved to session but DB error: {e}")
            else:
                st.success("✅ Customer Journey saved (session only)")

    elif mat_tools == "Mystery Shopping":
        st.subheader("Mystery Shopping")
        location = st.text_input("Location Evaluated")
        date = st.date_input("Date of Visit")
        col1, col2 = st.columns(2)
        with col1:
            ambiance = st.slider("Ambiance/Interface", 1, 10, 5)
            accessibility = st.slider("Accessibility", 1, 10, 5)
            product_display = st.slider("Product Display", 1, 10, 5)
        with col2:
            staff_behavior = st.slider("Staff Behavior", 1, 10, 5)
            product_knowledge = st.slider("Product Knowledge", 1, 10, 5)
            response_time = st.slider("Response Time", 1, 10, 5)
        observations = st.text_area("Detailed Observations")
        recommendations = st.text_area("Recommendations")

        if st.button("Save Mystery Shopping Report"):
            if 'mystery_shopping' not in st.session_state.mat_data:
                st.session_state.mat_data['mystery_shopping'] = []
            entry = {'location': location, 'date': str(date),
                     'ambiance': ambiance, 'accessibility': accessibility,
                     'product_display': product_display, 'staff_behavior': staff_behavior,
                     'product_knowledge': product_knowledge, 'response_time': response_time,
                     'observations': observations, 'recommendations': recommendations}
            st.session_state.mat_data['mystery_shopping'].append(entry)
            if st.session_state.current_project_id:
                try:
                    db.save_mat_mystery_shopping(_pid(), entry)
                    st.success("✅ Report saved to database")
                except Exception as e:
                    st.warning(f"Saved to session but DB error: {e}")
            else:
                st.success("✅ Report saved (session only)")

    elif mat_tools == "Complaint Data Analysis":
        st.subheader("Complaint Data Analysis")
        complaint_source = st.text_input("Complaint Source")
        complaint_date = st.date_input("Date Received")
        complaint_category = st.selectbox("Complaint Category",
                                          ["Product Quality", "Service", "Delivery", "Pricing", "Communication", "Other"])
        complaint_description = st.text_area("Complaint Description")
        severity = st.select_slider("Severity", options=["Low", "Medium", "High", "Critical"])
        resolution_status = st.selectbox("Resolution Status", ["Pending", "In Progress", "Resolved", "Closed"])
        resolution_details = st.text_area("Resolution Details")

        if st.button("Add Complaint"):
            if 'complaints' not in st.session_state.mat_data:
                st.session_state.mat_data['complaints'] = []
            entry = {'source': complaint_source, 'date': str(complaint_date),
                     'category': complaint_category, 'description': complaint_description,
                     'severity': severity, 'status': resolution_status, 'resolution': resolution_details}
            st.session_state.mat_data['complaints'].append(entry)
            if st.session_state.current_project_id:
                try:
                    db.save_mat_complaint(_pid(), entry)
                    st.success("✅ Complaint logged to database")
                except Exception as e:
                    st.warning(f"Saved to session but DB error: {e}")
            else:
                st.success("✅ Complaint logged (session only)")

        if 'complaints' in st.session_state.mat_data and st.session_state.mat_data['complaints']:
            df_complaints = pd.DataFrame(st.session_state.mat_data['complaints'])
            col1, col2 = st.columns(2)
            with col1:
                category_counts = df_complaints['category'].value_counts()
                fig_cat = px.pie(values=category_counts.values, names=category_counts.index, title="By Category")
                st.plotly_chart(fig_cat, use_container_width=True)
            with col2:
                severity_counts = df_complaints['severity'].value_counts()
                fig_sev = px.bar(x=severity_counts.index, y=severity_counts.values, title="By Severity")
                st.plotly_chart(fig_sev, use_container_width=True)

    elif mat_tools == "Brand Audit":
        st.subheader("Brand Audit")
        col1, col2 = st.columns(2)
        with col1:
            brand_mission = st.text_area("Brand Mission")
            brand_vision = st.text_area("Brand Vision")
            brand_values = st.text_area("Brand Values")
        with col2:
            usp = st.text_area("Unique Selling Proposition")
            brand_personality = st.text_area("Brand Personality")
            brand_promise = st.text_area("Brand Promise")

        col1, col2, col3 = st.columns(3)
        with col1:
            brand_awareness = st.slider("Brand Awareness", 1, 10, 5)
            brand_recognition = st.slider("Brand Recognition", 1, 10, 5)
        with col2:
            brand_loyalty = st.slider("Brand Loyalty", 1, 10, 5)
            customer_satisfaction = st.slider("Customer Satisfaction", 1, 10, 5)
        with col3:
            market_position = st.slider("Market Position", 1, 10, 5)
            brand_consistency = st.slider("Brand Consistency", 1, 10, 5)

        strengths = st.text_area("Brand Strengths")
        weaknesses = st.text_area("Brand Weaknesses")
        recommendations = st.text_area("Strategic Recommendations")

        if st.button("Save Brand Audit"):
            data = {'brand_mission': brand_mission, 'brand_vision': brand_vision,
                    'brand_values': brand_values, 'usp': usp, 'personality': brand_personality,
                    'promise': brand_promise, 'awareness': brand_awareness, 'recognition': brand_recognition,
                    'loyalty': brand_loyalty, 'satisfaction': customer_satisfaction,
                    'position': market_position, 'consistency': brand_consistency,
                    'strengths': strengths, 'weaknesses': weaknesses, 'recommendations': recommendations}
            st.session_state.mat_data['brand_audit'] = data
            if st.session_state.current_project_id:
                try:
                    db.save_mat_brand_audit(_pid(), data)
                    st.success("✅ Brand Audit saved to database")
                except Exception as e:
                    st.warning(f"Saved to session but DB error: {e}")
            else:
                st.success("✅ Brand Audit saved (session only)")


# ─────────────────────────────────────────────
# NATURE OF CRAFT
# ─────────────────────────────────────────────
elif _page == "NOC":
    st.markdown('<div class="main-header">Nature of Craft - 5P Framework</div>', unsafe_allow_html=True)
    st.markdown("""
    The 5P Framework helps define and analyse the nature of your craft across five key dimensions:
    **Product**, **Proficiency**, **Process**, **Purpose**, and **Portrayal**.
    """)

    framework_data = {
        "Product (Any Craft product)": {
            "secondary": ["Utilitarian", "Decorative Artefacts"],
            "tertiary": ["Functional Utility"]
        },
        "Proficiency (Needs a skill)": {
            "secondary": ["Material Understanding", "Practical knowledge", "Craftsmanship", "Experience", "Emotional Value Creation"],
            "tertiary": ["Skill-based Activity", "Intuitive Learning", "Skill sharing", "Craft Disciplines", "Sensory Experience", "Aesthetic Judgement", "Craft Knowledge"]
        },
        "Process (follows a process)": {
            "secondary": ["Sustainable manufacturing", "Material Manipulation Techniques", "Physical World Interaction", "Local Production", "Network Engagement"],
            "tertiary": ["Eco-effective process", "Non-industrial Production", "Community Collaboration method"]
        },
        "Purpose (to fulfil a purpose)": {
            "secondary": ["Prosumption", "Contextual lifestyle", "Economical aspect", "Self-satisfaction", "Aesthetic value"],
            "tertiary": ["Consumer Market Focus", "Consumer Behaviour", "Community Economy"]
        },
        "Portrayal (portrays a meaning)": {
            "secondary": ["Social Significance", "Creative Expression", "Cultural and Religious Representation"],
            "tertiary": ["Traditional Folkloric", "Ideology", "Local Culture", "Cultural Heritage", "Cultural Symbolism", "Traditional Wisdom", "Self-Expression", "Individual Works Conception"]
        }
    }

    status_tabs = st.tabs(["📍 Current Status", "🎯 Desired New Status"])

    for tab_idx, (tab_obj, status_type) in enumerate(zip(status_tabs, ['current', 'desired'])):
        with tab_obj:
            label = "Current Status" if status_type == 'current' else "Desired New Status"
            st.markdown(f"### {label} of Your Craft")
            status_key = f'{status_type}_status'

            for primary, content in framework_data.items():
                with st.expander(f"**{primary}**", expanded=False):
                    st.markdown("**Secondary Nature:**")
                    for item in content['secondary']:
                        key = f"{status_type}_{primary}_{item}"
                        if key not in st.session_state.nature_of_craft[status_key]:
                            st.session_state.nature_of_craft[status_key][key] = False
                        st.session_state.nature_of_craft[status_key][key] = st.checkbox(
                            item,
                            value=st.session_state.nature_of_craft[status_key][key],
                            key=f"cb_{key}"
                        )
                    if content['tertiary']:
                        st.markdown("**Tertiary Nature:**")
                        for item in content['tertiary']:
                            key = f"{status_type}_{primary}_{item}"
                            if key not in st.session_state.nature_of_craft[status_key]:
                                st.session_state.nature_of_craft[status_key][key] = False
                            st.session_state.nature_of_craft[status_key][key] = st.checkbox(
                                item,
                                value=st.session_state.nature_of_craft[status_key][key],
                                key=f"cb_{key}"
                            )

            if st.button(f"💾 Save {label}", use_container_width=True, type="primary", key=f"save_noc_{status_type}"):
                selections = st.session_state.nature_of_craft[status_key]
                count = sum(1 for v in selections.values() if v)
                if st.session_state.current_project_id:
                    try:
                        db.save_nature_of_craft(_pid(), status_type, selections)
                        st.success(f"✅ {label} saved to database ({count} characteristics selected)")
                    except Exception as e:
                        st.warning(f"Session saved but DB error: {e}")
                else:
                    st.success(f"✅ {label} saved (session only — {count} selected)")


# ─────────────────────────────────────────────
# DASHBOARD  (SIAMA native visual language)
# ─────────────────────────────────────────────
elif _page == "DASH":
    sit  = st.session_state.sit_data
    sat  = st.session_state.sat_data
    mat  = st.session_state.mat_data
    noc  = st.session_state.get("nature_of_craft", {"current_status": {}, "desired_status": {}})

    total_actors = sum(len(a) for a in sit.get("roles", {}).values())
    rel_data     = sat.get("relationship_data") or []
    n_ratings    = len(rel_data)
    n_subgroups  = len(sat.get("subgroups") or {})

    # ── Page heading — same .main-header class as every other page ──
    st.markdown('<div class="main-header">📊 Dashboard & Export</div>', unsafe_allow_html=True)

    # ── No-project guard ──
    if not st.session_state.get("current_project_id"):
        st.info("📁 Go to **Projects** in the sidebar to create or load a project. The dashboard will populate as you fill in SIT, SAT, MAT, and Nature of Craft.")
        st.stop()

    # ── Completion strip ──
    def _cs(pct):
        if pct == 100: return "cs-done"
        if pct > 0:    return "cs-active"
        return ""

    st.markdown(f"""
        <div class="completion-strip">
            <div class="cs-item cs-done">
                <div class="cs-name">Setup</div>
                <div class="cs-count">Done</div>
            </div>
            <div class="cs-item {_cs(sit_pct)}">
                <div class="cs-name">SIT</div>
                <div class="cs-count">{sit_done}/{sit_total} steps</div>
            </div>
            <div class="cs-item {_cs(sat_pct)}">
                <div class="cs-name">SAT</div>
                <div class="cs-count">{sat_done}/{sat_total} steps</div>
            </div>
            <div class="cs-item {_cs(mat_pct)}">
                <div class="cs-name">MAT</div>
                <div class="cs-count">{mat_done}/{mat_total} tools</div>
            </div>
            <div class="cs-item {_cs(noc_pct)}">
                <div class="cs-name">Nature</div>
                <div class="cs-count">{noc_done}/{noc_total}</div>
            </div>
            <div class="cs-item">
                <div class="cs-name">Export</div>
                <div class="cs-count">—</div>
            </div>
        </div>
    """, unsafe_allow_html=True)

    # ── KPI row ──
    if rel_data:
        df_all = pd.DataFrame(rel_data)
        critical = len(df_all[(df_all["power"] > 5) & (df_all["interest"] > 5)])
    else:
        critical = 0

    k1, k2, k3, k4, k5 = st.columns(5)
    kpi_data = [
        (k1, "blue",   "Actors Mapped",         total_actors,   f"SIT — {len([r for r,a in sit.get('roles',{}).items() if a])} roles"),
        (k2, "green",  "Stakeholders Rated",    n_ratings,      "in SAT matrix"),
        (k3, "orange", "Subgroups Formed",       n_subgroups,    "manual + auto"),
        (k4, "purple", "MAT Tools Done",         f"{mat_done}/8","market analysis"),
        (k5, "blue",   "Key Players",            critical,       "High P × High I"),
    ]
    for col, accent, label, val, sub in kpi_data:
        with col:
            st.markdown(f"""
                <div class="kpi-card {accent}">
                    <div class="kpi-label">{label}</div>
                    <div class="kpi-value">{val}</div>
                    <div class="kpi-sub">{sub}</div>
                </div>
            """, unsafe_allow_html=True)

    st.write("")

    # ── Row 1: Toolkit Progress + Priority + Actors ──
    col_left, col_mid, col_right = st.columns([1.1, 1.2, 1.2])

    with col_left:
        st.markdown("""
            <div class="dash-panel" style="height:auto;">
                <h4>📋 Toolkit Progress</h4>
                <div class="panel-sub">Steps completed per toolkit</div>
        """, unsafe_allow_html=True)

        for tk_name, done, total, pct, steps in [
            ("SIT — Stakeholder Identification", sit_done, sit_total, sit_pct,
             ["Questionnaires","Actor Database","Role Card","Role Map"]),
            ("SAT — Stakeholder Analysis", sat_done, sat_total, sat_pct,
             ["Relationship Matrix","Management Tool","Conflict Resolution","Knowledge Map","Value Exchange"]),
            ("MAT — Market Analysis", mat_done, mat_total, mat_pct,
             ["PESTEL","Gap","Segments","Personas","Journey","Mystery","Complaints","Brand"]),
            ("Nature of Craft", noc_done, noc_total, noc_pct,
             ["Current Status","Desired Status"]),
        ]:
            bar_cls = "done" if pct == 100 else ""
            pills = ""
            for i, s in enumerate(steps):
                if i < done:
                    pills += f'<span class="tpc-step-pill done">{s}</span>'
                else:
                    pills += f'<span class="tpc-step-pill">{s}</span>'

            st.markdown(f"""
                <div class="toolkit-progress-card">
                    <div class="tpc-header">
                        <div class="tpc-name">{tk_name}</div>
                        <div class="tpc-count">{done}/{total}</div>
                    </div>
                    <div class="tpc-bar">
                        <span style="width:{pct}%" class="{bar_cls}"></span>
                    </div>
                    <div class="tpc-steps">{pills}</div>
                </div>
            """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with col_mid:
        # Actors by role
        st.markdown("""
            <div class="dash-panel" style="margin-bottom:12px;">
                <h4>👥 Actors by Role</h4>
                <div class="panel-sub">SIT — stakeholder identification</div>
        """, unsafe_allow_html=True)
        if sit.get("roles"):
            role_counts = {r: len(a) for r, a in sit["roles"].items() if len(a) > 0}
            mx = max(role_counts.values()) if role_counts else 1
            rows = ""
            for role, cnt in role_counts.items():
                pct = int(cnt / mx * 100)
                rows += f"""
                    <div class="role-bar-row">
                        <div class="rbl-name">{role}</div>
                        <div class="rbl-track"><span style="width:{pct}%"></span></div>
                        <div class="rbl-count">{cnt}</div>
                    </div>"""
            st.markdown(rows + "</div>", unsafe_allow_html=True)
        else:
            st.markdown("<p style='color:#6c757d;font-size:13px;'>Add actors in SIT → Step 2</p></div>", unsafe_allow_html=True)

        # Quadrant distribution
        st.markdown("""
            <div class="dash-panel">
                <h4>🎯 Engagement Quadrants</h4>
                <div class="panel-sub">SAT — stakeholder management strategy</div>
        """, unsafe_allow_html=True)
        if rel_data:
            df = pd.DataFrame(rel_data)
            qc = {"key": 0, "satisfy": 0, "inform": 0, "monitor": 0}
            for _, r in df.iterrows():
                hp, hi = r["power"] > 5, r["interest"] > 5
                if hp and hi:       qc["key"] += 1
                elif hp and not hi: qc["satisfy"] += 1
                elif not hp and hi: qc["inform"] += 1
                else:               qc["monitor"] += 1
            st.markdown(f"""
                <div class="quad-grid">
                    <div class="quad-tile qt-manage">
                        <div class="qt-strategy">Manage Closely</div>
                        <div class="qt-axis">High Power · High Interest</div>
                        <div class="qt-count">{qc["key"]}</div>
                    </div>
                    <div class="quad-tile qt-satisfy">
                        <div class="qt-strategy">Keep Satisfied</div>
                        <div class="qt-axis">High Power · Low Interest</div>
                        <div class="qt-count">{qc["satisfy"]}</div>
                    </div>
                    <div class="quad-tile qt-inform">
                        <div class="qt-strategy">Keep Informed</div>
                        <div class="qt-axis">Low Power · High Interest</div>
                        <div class="qt-count">{qc["inform"]}</div>
                    </div>
                    <div class="quad-tile qt-monitor">
                        <div class="qt-strategy">Monitor</div>
                        <div class="qt-axis">Low Power · Low Interest</div>
                        <div class="qt-count">{qc["monitor"]}</div>
                    </div>
                </div>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown("<p style='color:#6c757d;font-size:13px;'>Rate stakeholders in SAT → Step 1</p></div>", unsafe_allow_html=True)

    with col_right:
        # Power vs Interest scatter
        st.markdown("""
            <div class="dash-panel" style="margin-bottom:12px;">
                <h4>📊 Power × Interest Map</h4>
                <div class="panel-sub">SAT — stakeholder positioning</div>
            </div>
        """, unsafe_allow_html=True)
        if rel_data:
            df = pd.DataFrame(rel_data)
            df["priority_score"] = df["power"].astype(float) * df["interest"].astype(float)
            def _qcolor(r):
                if r["power"] > 5 and r["interest"] > 5: return "#1f77b4"
                if r["power"] > 5: return "#fd7e14"
                if r["interest"] > 5: return "#28a745"
                return "#6c757d"
            df["_c"] = df.apply(_qcolor, axis=1)
            sc = go.Figure()
            sc.add_trace(go.Scatter(
                x=df["power"], y=df["interest"],
                mode="markers+text",
                text=df["stakeholder"].astype(str).str.split("(").str[0].str.strip(),
                textposition="top center",
                textfont=dict(size=9, color="#495057"),
                marker=dict(
                    size=df["priority_score"] * 0.5 + 10,
                    color=df["_c"],
                    line=dict(color="#ffffff", width=1.5),
                    opacity=0.9,
                ),
                hovertext=[f"{s}<br>Power:{p}  Interest:{i}" for s,p,i in zip(df["stakeholder"],df["power"],df["interest"])],
                hoverinfo="text",
            ))
            sc.add_hline(y=5, line_dash="dash", line_color="#dee2e6", line_width=1)
            sc.add_vline(x=5, line_dash="dash", line_color="#dee2e6", line_width=1)
            sc.update_layout(
                height=260,
                margin=dict(l=30, r=10, t=10, b=30),
                xaxis=dict(range=[0,11], title="Power", gridcolor="#f0f0f0", zeroline=False,
                           title_font=dict(size=10,color="#6c757d")),
                yaxis=dict(range=[0,11], title="Interest", gridcolor="#f0f0f0", zeroline=False,
                           title_font=dict(size=10,color="#6c757d")),
                plot_bgcolor="#ffffff",
                paper_bgcolor="rgba(0,0,0,0)",
                showlegend=False,
            )
            st.plotly_chart(sc, use_container_width=True)
        else:
            st.caption("Rate stakeholders in SAT → Step 1 to see positioning.")

        # Priority list
        st.markdown("""
            <div class="dash-panel">
                <h4>🏆 Top Stakeholders</h4>
                <div class="panel-sub">Ranked by Power × Interest score</div>
        """, unsafe_allow_html=True)
        if rel_data:
            df = pd.DataFrame(rel_data).copy()
            df["score"] = df["power"].astype(float) * df["interest"].astype(float)
            top5 = df.sort_values("score", ascending=False).head(5)
            items = ""
            for i, (_, r) in enumerate(top5.iterrows(), 1):
                sh = str(r["stakeholder"])
                name = sh.rsplit("(", 1)[0].strip() if "(" in sh else sh
                role = sh.rsplit("(", 1)[1].rstrip(")") if "(" in sh else ""
                items += f"""
                    <div class="pri-list-item">
                        <div class="pri-rank">{i}</div>
                        <div class="pri-info">
                            <div class="pname">{name}</div>
                            {"" if not role else f'<div class="prole">{role}</div>'}
                        </div>
                        <div class="pri-score">{int(r["score"])}</div>
                    </div>"""
            st.markdown(items + "</div>", unsafe_allow_html=True)
        else:
            st.markdown("<p style='color:#6c757d;font-size:13px;'>No ratings yet.</p></div>", unsafe_allow_html=True)

    st.write("")

    # ── Row 2: MAT Coverage + 5P Nature of Craft ──
    mc1, mc2 = st.columns([1.5, 1])

    with mc1:
        st.markdown("""
            <div class="dash-panel">
                <h4>🔬 MAT Toolkit Coverage</h4>
                <div class="panel-sub">Market Analysis Tools — 8 tools total</div>
        """, unsafe_allow_html=True)
        tool_specs = [
            ("pestel", "PESTEL", "1"),("gap","Gap Analysis","2"),
            ("behavioral_segments","Behavioural Segs","3"),("personas","Personas","4"),
            ("customer_journey","Customer Journey","5"),("mystery_shopping","Mystery Shopping","6"),
            ("complaints","Complaints","7"),("brand_audit","Brand Audit","8"),
        ]
        cells = ""
        for key, label, num in tool_specs:
            done_cls = "mat-done" if bool(mat.get(key)) else ""
            check = "✓" if done_cls else ""
            cells += f"""
                <div class="mat-tile {done_cls}">
                    <div class="mt-num">{num} {check}</div>
                    <div class="mt-name">{label}</div>
                </div>"""
        st.markdown(f'<div class="mat-coverage-grid">{cells}</div></div>', unsafe_allow_html=True)

    with mc2:
        st.markdown("""
            <div class="dash-panel">
                <h4>🌿 Nature of Craft — 5P</h4>
                <div class="panel-sub">
                    <span style="color:#1f77b4">■</span> Current &nbsp;
                    <span style="color:#fd7e14">■</span> Desired
                </div>
        """, unsafe_allow_html=True)

        five_ps = [("Product","P"),("Production","P"),("Process","P"),("Purpose","P"),("Portrayal","P")]
        cur = noc.get("current_status", {}) or {}
        des = noc.get("desired_status", {}) or {}

        def _count_p(d, name):
            return sum(1 for k, v in d.items() if v and f"_{name}" in k)

        all_cur = [_count_p(cur, p) for p, _ in five_ps]
        all_des = [_count_p(des, p) for p, _ in five_ps]
        mx = max(all_cur + all_des + [1])

        rows = ""
        for (pname, letter), c, d in zip(five_ps, all_cur, all_des):
            cp = int(c / mx * 100)
            dp = int(d / mx * 100)
            rows += f"""
                <div class="fivep-row">
                    <div class="fp-p">{letter}</div>
                    <div class="fp-name">{pname}</div>
                    <div style="flex:1; display:flex; flex-direction:column; gap:2px;">
                        <div class="fp-track"><span class="cur" style="width:{cp}%"></span></div>
                        <div class="fp-track"><span class="des" style="width:{dp}%"></span></div>
                    </div>
                    <div class="fp-cnt">{c}/{d}</div>
                </div>"""
        if not any(all_cur) and not any(all_des):
            rows = "<p style='font-size:12px;color:#6c757d;margin:8px 0 0;'>Fill in Nature of Craft to populate this panel.</p>"
        st.markdown(rows + "</div>", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# SUMMARY & EXPORT
# ─────────────────────────────────────────────

    # ─── EXPORT & TRAINING RECOMMENDATIONS (merged from Summary page) ───
    st.markdown("---")
    st.subheader("📥 Export Data")
    _export_format = st.radio("Select Export Format", ["JSON", "Excel"], key="export_fmt_dash")

    if st.button("Generate Export File", key="export_btn_dash"):
        if _export_format == "JSON":
            _export_data = {
                "project": st.session_state.current_project,
                "sit_data": st.session_state.sit_data,
                "sat_data": st.session_state.sat_data,
                "mat_data": st.session_state.mat_data,
                "export_date": datetime.now().isoformat()
            }
            st.download_button(
                label="📥 Download JSON",
                data=json.dumps(_export_data, indent=2, default=_json_default),
                file_name=f"siama_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
        elif _export_format == "Excel":
            try:
                import io
                from openpyxl import Workbook
                from openpyxl.styles import Font
                wb = Workbook()
                ws_sit = wb.active
                ws_sit.title = "SIT Data"
                ws_sit["A1"] = "Stakeholder Identification Data"
                ws_sit["A1"].font = Font(bold=True, size=14)
                _row = 3
                ws_sit[f"A{_row}"] = "Role"; ws_sit[f"B{_row}"] = "Actor Name"
                ws_sit[f"C{_row}"] = "Location"; ws_sit[f"D{_row}"] = "Contact"
                _row += 1
                for _role, _actors in st.session_state.sit_data.get("roles", {}).items():
                    for _actor in _actors:
                        ws_sit[f"A{_row}"] = _role
                        ws_sit[f"B{_row}"] = _actor.get("name", "")
                        ws_sit[f"C{_row}"] = _actor.get("location", "")
                        ws_sit[f"D{_row}"] = _actor.get("contact", "")
                        _row += 1
                if "relationship_data" in st.session_state.sat_data:
                    ws_sat = wb.create_sheet("SAT Data")
                    ws_sat["A1"] = "Stakeholder Analysis Data"
                    ws_sat["A1"].font = Font(bold=True, size=14)
                    _df_sat = pd.DataFrame(st.session_state.sat_data["relationship_data"])
                    for _r_idx, _row_data in enumerate(_df_sat.values, start=3):
                        for _c_idx, _value in enumerate(_row_data, start=1):
                            ws_sat.cell(row=_r_idx, column=_c_idx, value=str(_value))
                _excel_buf = io.BytesIO()
                wb.save(_excel_buf); _excel_buf.seek(0)
                st.download_button(
                    label="📥 Download Excel",
                    data=_excel_buf,
                    file_name=f"siama_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Error creating Excel file: {e}")

    st.markdown("---")
    st.subheader("🎯 Training Recommendations")
    if st.button("Generate Recommendations", key="reco_btn_dash"):
        if st.session_state.sit_data.get("roles"):
            st.markdown("**Stakeholder Landscape:**")
            for _role, _actors in st.session_state.sit_data["roles"].items():
                st.write(f"- {_role}: {len(_actors)} actors identified")
        if "relationship_data" in st.session_state.sat_data and st.session_state.sat_data["relationship_data"]:
            _df_r = pd.DataFrame(st.session_state.sat_data["relationship_data"])
            _hp = len(_df_r[(_df_r["power"] > 5) & (_df_r["interest"] > 5)])
            st.write(f"- {_hp} key stakeholders requiring close management")
        if st.session_state.mat_data:
            st.write(f"- {len(st.session_state.mat_data)} market analysis tools completed")
        st.markdown("""
        #### Recommended Training Focus Areas:
        1. **Technical Skills Development** — Based on identified producer gaps
        2. **Business & Entrepreneurship** — Market understanding, pricing, financial management
        3. **Digital Literacy** — E-commerce, social media, digital payments
        4. **Stakeholder Collaboration** — Supplier relations, customer management, networking
        5. **Design & Innovation** — Market trends, product diversification, cultural authenticity
        """)

# FOOTER
# ─────────────────────────────────────────────
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: gray;'>
    <p>SIAMA Toolbox v2.0 | Developed for Craft Education Planning in India</p>
    <p>Based on research by Kumar et al., IIT Guwahati | Database: Supabase (PostgreSQL)</p>
</div>
""", unsafe_allow_html=True)
